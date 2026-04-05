import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Colors
NAVY = "0A1628"
GOLD = "BFA046"
LIGHT_GOLD = "F5EDD6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F4"
MEDIUM_GRAY = "D9D9D9"
DARK_TEXT = "1A1A1A"
GREEN = "2E7D32"

# Styles
header_font = Font(name="Arial", size=12, bold=True, color=GOLD)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

subheader_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")

input_font = Font(name="Arial", size=11, color=DARK_TEXT)
input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

result_font = Font(name="Arial", size=11, bold=True, color=NAVY)
result_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

normal_font = Font(name="Arial", size=11, color=DARK_TEXT)
bold_font = Font(name="Arial", size=11, bold=True, color=DARK_TEXT)
title_font = Font(name="Arial", size=16, bold=True, color=GOLD)
title_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

row_even = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
row_odd = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

nis_format = '₪#,##0'
nis_format_dec = '₪#,##0.00'
pct_format = '0.0%'

VAT_RATE = 0.18

# Supplier costs (per sqm, before VAT)
MIRROR_GLASS_PER_SQM = 197
UV_PRINT_PER_SQM = 380
SPRAY_PER_SQM = 100
MIRROR_TOTAL_PER_SQM = MIRROR_GLASS_PER_SQM + UV_PRINT_PER_SQM + SPRAY_PER_SQM  # 677

LED_COST = 30
LED_MARKUP = 200  # sells at cost + 200

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: מחשבון תמחור (Pricing Calculator)
# ============================================================
ws1 = wb.active
ws1.title = "מחשבון תמחור"
ws1.sheet_view.rightToLeft = True

# Column widths
ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 26

# Title
ws1.merge_cells('B1:D1')
ws1['B1'] = "Mirror Frame Studio - מחשבון תמחור"
ws1['B1'].font = title_font
ws1['B1'].fill = title_fill
ws1['B1'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 5):
    ws1.cell(row=1, column=c).fill = title_fill
    ws1.cell(row=1, column=c).border = thin_border
ws1.row_dimensions[1].height = 40

# --- INPUT SECTION ---
ws1.merge_cells('B3:D3')
ws1['B3'] = "נתוני הזמנה"
ws1['B3'].font = Font(name="Arial", size=13, bold=True, color=NAVY)
ws1['B3'].fill = subheader_fill
ws1['B3'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 5):
    ws1.cell(row=3, column=c).fill = subheader_fill
    ws1.cell(row=3, column=c).border = thin_border

inputs = [
    (5, "רוחב המראה (ס\"מ)", 80),
    (6, "גובה המראה (ס\"מ)", 160),
    (7, "LED", "כן"),
    (8, "כמות", 1),
]

for row_num, label, default in inputs:
    ws1.cell(row=row_num, column=2, value=label)
    ws1.cell(row=row_num, column=2).font = bold_font
    ws1.cell(row=row_num, column=2).alignment = Alignment(horizontal="right", vertical="center")
    ws1.cell(row=row_num, column=2).border = thin_border

    ws1.cell(row=row_num, column=3, value=default)
    ws1.cell(row=row_num, column=3).font = input_font
    ws1.cell(row=row_num, column=3).fill = input_fill
    ws1.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row_num, column=3).border = thin_border

# LED yes/no validation
dv_yn = DataValidation(type="list", formula1='"כן,לא"', allow_blank=False)
ws1.add_data_validation(dv_yn)
dv_yn.add('C7')

# --- COST BREAKDOWN ---
ws1.merge_cells('B10:D10')
ws1['B10'] = "פירוט עלויות (לפני מע\"מ)"
ws1['B10'].font = Font(name="Arial", size=13, bold=True, color=NAVY)
ws1['B10'].fill = subheader_fill
ws1['B10'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 5):
    ws1.cell(row=10, column=c).fill = subheader_fill
    ws1.cell(row=10, column=c).border = thin_border

calc_rows = [
    (11, "שטח המראה (מ\"ר)", '=C5*C6/10000', '0.000', None),
    (12, "זכוכית מראה 6 מ\"מ", '=C11*197', nis_format_dec, "₪197/מ\"ר"),
    (13, "הדפסה דיגיטלית UV", '=C11*380', nis_format_dec, "₪380/מ\"ר"),
    (14, "התזה", '=C11*100', nis_format_dec, "₪100/מ\"ר"),
    (15, "סה\"כ מראה מספק", '=C12+C13+C14', nis_format_dec, "₪677/מ\"ר"),
    (16, "עלות LED", '=IF(C7="כן",30,0)', nis_format_dec, "₪30"),
]

for row_num, label, formula, fmt, note in calc_rows:
    ws1.cell(row=row_num, column=2, value=label)
    ws1.cell(row=row_num, column=2).font = normal_font
    ws1.cell(row=row_num, column=2).alignment = Alignment(horizontal="right", vertical="center")
    ws1.cell(row=row_num, column=2).border = thin_border

    ws1.cell(row=row_num, column=3, value=formula)
    ws1.cell(row=row_num, column=3).font = normal_font
    ws1.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row_num, column=3).border = thin_border
    ws1.cell(row=row_num, column=3).number_format = fmt

    if note:
        ws1.cell(row=row_num, column=4, value=note)
        ws1.cell(row=row_num, column=4).font = Font(name="Arial", size=9, italic=True, color="888888")
        ws1.cell(row=row_num, column=4).alignment = Alignment(horizontal="center", vertical="center")

    fill = row_even if (row_num - 11) % 2 == 0 else row_odd
    ws1.cell(row=row_num, column=2).fill = fill
    ws1.cell(row=row_num, column=3).fill = fill

# Bold the subtotal row
ws1.cell(row=15, column=2).font = bold_font
ws1.cell(row=15, column=3).font = bold_font

# --- RESULTS SECTION ---
ws1.merge_cells('B18:D18')
ws1['B18'] = "סיכום תמחור"
ws1['B18'].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws1['B18'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws1['B18'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 5):
    ws1.cell(row=18, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws1.cell(row=18, column=c).border = thin_border

results = [
    (19, "עלות ייצור ליחידה (לפני מע\"מ)", '=C15+C16', nis_format),
    (20, "עלות כולל מע\"מ", '=C19*1.18', nis_format),
    (21, "מחיר מכירה ללקוח + מע\"מ", '=CEILING(C19*2,100)', nis_format),
    (22, "רווח ליחידה (לפני מע\"מ)", '=C21-C19', nis_format),
    (23, "סה\"כ עלות להזמנה", '=C20*C8', nis_format),
    (24, "סה\"כ הכנסה מהזמנה", '=C21*1.18*C8', nis_format),
    (25, "סה\"כ רווח מהזמנה", '=C24-C23', nis_format),
]

for row_num, label, formula, fmt in results:
    ws1.cell(row=row_num, column=2, value=label)
    ws1.cell(row=row_num, column=2).font = bold_font
    ws1.cell(row=row_num, column=2).alignment = Alignment(horizontal="right", vertical="center")
    ws1.cell(row=row_num, column=2).border = thin_border

    ws1.cell(row=row_num, column=3, value=formula)
    ws1.cell(row=row_num, column=3).font = result_font
    ws1.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row_num, column=3).border = thin_border
    ws1.cell(row=row_num, column=3).number_format = fmt

    ws1.cell(row=row_num, column=2).fill = result_fill
    ws1.cell(row=row_num, column=3).fill = result_fill

# Highlight sell price and total profit
for r in [21, 25]:
    ws1.cell(row=r, column=2).font = Font(name="Arial", size=12, bold=True, color=GREEN)
    ws1.cell(row=r, column=3).font = Font(name="Arial", size=13, bold=True, color=GREEN)
    ws1.cell(row=r, column=2).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws1.cell(row=r, column=3).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

# Note
ws1.merge_cells('B29:D29')
ws1['B29'] = "* מחירי ספק: א.א. מראות (הצעה 23102116, מרץ 2026) | לפני מע\"מ"
ws1['B29'].font = Font(name="Arial", size=9, italic=True, color="777777")
ws1['B29'].alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# SHEET 2: מחירון (Price List)
# ============================================================
ws2 = wb.create_sheet("מחירון")
ws2.sheet_view.rightToLeft = True

for col_letter, w in [('A', 4), ('B', 18), ('C', 16), ('D', 14),
                       ('E', 16), ('F', 32)]:
    ws2.column_dimensions[col_letter].width = w

# Title
ws2.merge_cells('B1:F1')
ws2['B1'] = "Mirror Frame Studio - מחירון"
ws2['B1'].font = title_font
ws2['B1'].fill = title_fill
ws2['B1'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 7):
    ws2.cell(row=1, column=c).fill = title_fill
    ws2.cell(row=1, column=c).border = thin_border
ws2.row_dimensions[1].height = 40

# Headers
headers2 = ["", "גודל", "מידות (ס\"מ)", "שטח (מ\"ר)",
            "עלות מספק", "מחיר ללקוח + מע\"מ"]
for i, h in enumerate(headers2):
    cell = ws2.cell(row=3, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws2.row_dimensions[3].height = 30

sizes = [
    ("קטן", "40x60", 0.24),
    ("בינוני", "60x80", 0.48),
    ("גדול", "80x120", 0.96),
    ("80x160", "80x160", 1.28),
    ("XL", "100x150", 1.5),
]

row = 4
for i, (size_name, dims, area) in enumerate(sizes):
    mirror_cost = area * MIRROR_TOTAL_PER_SQM + LED_COST
    sell_before_vat = math.ceil(mirror_cost * 2 / 100) * 100  # round up to nearest 100

    ws2.cell(row=row, column=2, value=size_name)
    ws2.cell(row=row, column=3, value=dims)
    ws2.cell(row=row, column=4, value=round(area, 2))
    ws2.cell(row=row, column=5, value=round(mirror_cost, 0))
    ws2.cell(row=row, column=5).number_format = nis_format
    sell_with_vat = round(sell_before_vat * 1.18)
    ws2.cell(row=row, column=6, value=f"₪{sell_before_vat:,} + מע\"מ = ₪{sell_with_vat:,}")

    fill = row_even if i % 2 == 0 else row_odd
    for c in range(1, 7):
        ws2.cell(row=row, column=c).fill = fill
        ws2.cell(row=row, column=c).border = thin_border
        ws2.cell(row=row, column=c).font = normal_font
        ws2.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")

    row += 1

# Note
row += 1
ws2.merge_cells(f'B{row}:F{row}')
ws2.cell(row=row, column=2, value="* ספק: א.א. מראות ₪677/מ\"ר | LED: עלות ₪30 | מע\"מ 18% | מכירה: x2 מעלות כולל מע\"מ | מרץ 2026")
ws2.cell(row=row, column=2).font = Font(name="Arial", size=9, italic=True, color="777777")
ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# SHEET 3: סיכום חודשי (Monthly Summary)
# ============================================================
ws3 = wb.create_sheet("סיכום חודשי")
ws3.sheet_view.rightToLeft = True

for col_letter, w in [('A', 4), ('B', 18), ('C', 16), ('D', 18), ('E', 18), ('F', 18), ('G', 16)]:
    ws3.column_dimensions[col_letter].width = w

ws3.merge_cells('B1:G1')
ws3['B1'] = "Mirror Frame Studio - סיכום חודשי 2026"
ws3['B1'].font = title_font
ws3['B1'].fill = title_fill
ws3['B1'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 8):
    ws3.cell(row=1, column=c).fill = title_fill
    ws3.cell(row=1, column=c).border = thin_border
ws3.row_dimensions[1].height = 40

headers3 = ["", "חודש", "מספר הזמנות", "סה\"כ הכנסות", "סה\"כ עלויות", "רווח", "אחוז רווח"]
for i, h in enumerate(headers3):
    cell = ws3.cell(row=3, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws3.row_dimensions[3].height = 30

months_heb = [
    "ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
    "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר"
]

for i, month in enumerate(months_heb):
    r = 4 + i
    ws3.cell(row=r, column=2, value=month)
    ws3.cell(row=r, column=2).font = bold_font
    ws3.cell(row=r, column=3).number_format = '#,##0'
    ws3.cell(row=r, column=4).number_format = nis_format
    ws3.cell(row=r, column=5).number_format = nis_format
    ws3.cell(row=r, column=6, value=f'=IF(AND(D{r}<>"",E{r}<>""),D{r}-E{r},"")')
    ws3.cell(row=r, column=6).number_format = nis_format
    ws3.cell(row=r, column=7, value=f'=IF(AND(D{r}<>"",D{r}<>0),F{r}/D{r},"")')
    ws3.cell(row=r, column=7).number_format = pct_format

    fill = row_even if i % 2 == 0 else row_odd
    for c in range(1, 8):
        ws3.cell(row=r, column=c).fill = fill
        ws3.cell(row=r, column=c).border = thin_border
        ws3.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

# Summary row
summary_row = 16
ws3.cell(row=summary_row, column=2, value="סה\"כ")
ws3.cell(row=summary_row, column=3, value='=SUM(C4:C15)')
ws3.cell(row=summary_row, column=3).number_format = '#,##0'
ws3.cell(row=summary_row, column=4, value='=SUM(D4:D15)')
ws3.cell(row=summary_row, column=4).number_format = nis_format
ws3.cell(row=summary_row, column=5, value='=SUM(E4:E15)')
ws3.cell(row=summary_row, column=5).number_format = nis_format
ws3.cell(row=summary_row, column=6, value='=SUM(F4:F15)')
ws3.cell(row=summary_row, column=6).number_format = nis_format
ws3.cell(row=summary_row, column=7, value='=IF(AND(D16<>"",D16<>0),F16/D16,"")')
ws3.cell(row=summary_row, column=7).number_format = pct_format

for c in range(1, 8):
    ws3.cell(row=summary_row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws3.cell(row=summary_row, column=c).font = Font(name="Arial", size=11, bold=True, color=GOLD)
    ws3.cell(row=summary_row, column=c).border = thin_border
    ws3.cell(row=summary_row, column=c).alignment = Alignment(horizontal="center", vertical="center")

# ============================================================
# SHEET 4: מעקב הזמנות (Order Tracking)
# ============================================================
ws4 = wb.create_sheet("מעקב הזמנות")
ws4.sheet_view.rightToLeft = True

for col_letter, w in [('A', 6), ('B', 14), ('C', 20), ('D', 12),
                       ('E', 12), ('F', 8), ('G', 16), ('H', 20),
                       ('I', 14)]:
    ws4.column_dimensions[col_letter].width = w

# Title
ws4.merge_cells('A1:I1')
ws4['A1'] = "Mirror Frame Studio - מעקב הזמנות"
ws4['A1'].font = title_font
ws4['A1'].fill = title_fill
ws4['A1'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(1, 10):
    ws4.cell(row=1, column=c).fill = title_fill
    ws4.cell(row=1, column=c).border = thin_border
ws4.row_dimensions[1].height = 40

# Headers
headers4 = ["#", "תאריך", "שם לקוח", "רוחב", "גובה", "LED",
            "עלות שלי", "מחיר ללקוח + מע\"מ", "סטטוס"]
for i, h in enumerate(headers4):
    cell = ws4.cell(row=3, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws4.row_dimensions[3].height = 30

# Status validation
dv_status = DataValidation(type="list", formula1='"ממתין,בייצור,מוכן,נשלח,בוטל"', allow_blank=True)
ws4.add_data_validation(dv_status)

# LED validation for orders
dv_led = DataValidation(type="list", formula1='"כן,לא"', allow_blank=True)
ws4.add_data_validation(dv_led)

# 50 empty rows for orders with formulas
for i in range(50):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i + 1)
    ws4.cell(row=r, column=1).font = normal_font
    ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Date column
    ws4.cell(row=r, column=2).number_format = 'DD/MM/YYYY'

    # Cost formula: area * 677 + LED
    ws4.cell(row=r, column=7,
             value=f'=IF(AND(D{r}<>"",E{r}<>""),D{r}*E{r}/10000*677+IF(F{r}="כן",30,0),"")')
    ws4.cell(row=r, column=7).number_format = nis_format

    # Sell price: CEILING(cost*2, 100) shown as "X + מע״מ"
    ws4.cell(row=r, column=8,
             value=f'=IF(G{r}<>"",CEILING(G{r}*2,100)&" + מע""מ","")')

    dv_status.add(f'I{r}')
    dv_led.add(f'F{r}')

    fill = row_even if i % 2 == 0 else row_odd
    for c in range(1, 10):
        ws4.cell(row=r, column=c).fill = fill
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
        if c in [2, 3, 4, 5, 6, 9]:  # input columns
            ws4.cell(row=r, column=c).fill = input_fill


# ============================================================
# SHEET 5: מחירון גדלים מותאמים (Custom Sizes)
# ============================================================
ws5 = wb.create_sheet("מחירון מותאם")
ws5.sheet_view.rightToLeft = True

for col_letter, w in [('A', 4), ('B', 12), ('C', 12), ('D', 14),
                       ('E', 16), ('F', 24)]:
    ws5.column_dimensions[col_letter].width = w

# Title
ws5.merge_cells('B1:F1')
ws5['B1'] = "Mirror Frame Studio - מחירון גדלים מותאמים"
ws5['B1'].font = title_font
ws5['B1'].fill = title_fill
ws5['B1'].alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 7):
    ws5.cell(row=1, column=c).fill = title_fill
    ws5.cell(row=1, column=c).border = thin_border
ws5.row_dimensions[1].height = 40

# Instruction
ws5.merge_cells('B2:F2')
ws5['B2'] = "הזן רוחב וגובה - המחיר מחושב אוטומטית (כולל LED)"
ws5['B2'].font = Font(name="Arial", size=10, italic=True, color="555555")
ws5['B2'].alignment = Alignment(horizontal="center", vertical="center")

# Headers
headers5 = ["", "רוחב (ס\"מ)", "גובה (ס\"מ)", "שטח (מ\"ר)",
            "עלות שלי", "מחיר ללקוח + מע\"מ"]
for i, h in enumerate(headers5):
    cell = ws5.cell(row=4, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws5.row_dimensions[4].height = 30

# 20 rows for custom sizes
for i in range(20):
    r = 5 + i
    # Area formula
    ws5.cell(row=r, column=4,
             value=f'=IF(AND(B{r}<>"",C{r}<>""),B{r}*C{r}/10000,"")')
    ws5.cell(row=r, column=4).number_format = '0.00'

    # Cost formula: area * 677 + LED(30)
    ws5.cell(row=r, column=5,
             value=f'=IF(D{r}<>"",D{r}*677+30,"")')
    ws5.cell(row=r, column=5).number_format = nis_format

    # Sell price: CEILING(cost*2, 100) + מע"מ
    ws5.cell(row=r, column=6,
             value=f'=IF(E{r}<>"",CEILING(E{r}*2,100)&" + מע""מ = ₪"&TEXT(CEILING(E{r}*2,100)*1.18,"#,##0"),"")')

    fill = row_even if i % 2 == 0 else row_odd
    for c in range(1, 7):
        ws5.cell(row=r, column=c).fill = fill
        ws5.cell(row=r, column=c).border = thin_border
        ws5.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
        if c in [2, 3]:  # input columns
            ws5.cell(row=r, column=c).fill = input_fill

# Pre-fill popular sizes
popular = [(40, 60), (50, 70), (60, 80), (70, 100), (80, 120), (80, 160), (100, 150), (100, 200)]
for i, (w, h) in enumerate(popular):
    r = 5 + i
    ws5.cell(row=r, column=2, value=w)
    ws5.cell(row=r, column=3, value=h)

# Note
ws5.merge_cells('B26:F26')
ws5['B26'] = "* עלות ספק: ₪677/מ\"ר + LED ₪30 | מכירה: x2 עיגול ל-100 + מע\"מ 18%"
ws5['B26'].font = Font(name="Arial", size=9, italic=True, color="777777")
ws5['B26'].alignment = Alignment(horizontal="center", vertical="center")


# Save
output_path = r"C:\Users\amith\ads-dashboard\mirror-frame-studio\Mirror_Frame_Studio_Pricing_Calculator.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
