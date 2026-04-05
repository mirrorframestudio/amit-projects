"""
תקציב משק בית - יומן הוצאות + הוצאות קבועות + סיכום חודשי + דשבורד
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, BarChart, Reference

wb = openpyxl.Workbook()

# ============================================================
# Shared Styles
# ============================================================
MONTHS_HEB = [
    "ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
    "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר"
]

CATEGORIES = [
    "מזון וסופר",
    "אכילה בחוץ / משלוחים",
    "תחבורה / דלק",
    "קניות / ביגוד",
    "בריאות / תרופות",
    "בילויים / פנאי",
    "חינוך / ילדים",
    "בית / תחזוקה",
    "חשבונות (חשמל/מים/גז)",
    "טלפון / אינטרנט",
    "מתנות",
    "חיות מחמד",
    "שונות",
]

PAYMENT_METHODS = [
    "מזומן",
    "אשראי",
    "העברה בנקאית",
    "ביט / פייבוקס",
    "צ'ק",
]

SPOUSES = [
    "בן/בת זוג 1",
    "בן/בת זוג 2",
    "משותף",
]

header_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
category_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
sub_header_font = Font(name="Arial", bold=True, size=11)
normal_font = Font(name="Arial", size=11)
total_font = Font(name="Arial", bold=True, size=12, color="1F4E79")
grand_total_font = Font(name="Arial", bold=True, size=13, color="FFFFFF")

title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
income_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
subtotal_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
header_gray = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
summary_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill_light = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
orange_fill_light = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
dark_red_fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
dashboard_fill = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
card_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

rtl_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
currency_format = '#,##0 ₪'


def style_cell(cell, font=normal_font, fill=None, align=center_align, fmt=None):
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


# ============================================================
# SHEET 1: יומן הוצאות (Daily Log) — now with "מי שילם" column
# ============================================================
ws_daily = wb.active
ws_daily.title = "יומן הוצאות"
ws_daily.sheet_view.rightToLeft = True

# Title
row = 1
ws_daily.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
cell = ws_daily.cell(row=row, column=1, value="יומן הוצאות יומי")
style_cell(cell, font=Font(name="Arial", bold=True, size=18, color="FFFFFF"), fill=title_fill)
ws_daily.row_dimensions[row].height = 40
row += 1

# Subtitle
ws_daily.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
cell = ws_daily.cell(row=row, column=1,
                     value="הוסיפו שורה חדשה לכל הוצאה - פשוט מלאו תאריך, סכום, קטגוריה ותיאור")
style_cell(cell, font=Font(name="Arial", size=10, italic=True, color="757575"),
           align=Alignment(horizontal="right", vertical="center"))
ws_daily.row_dimensions[row].height = 22
row += 1

# Headers — added "מי שילם" (column 5)
daily_headers = ["תאריך", "סכום", "קטגוריה", "אמצעי תשלום", "מי שילם", "תיאור / פירוט", "הערות"]
col_widths = [14, 14, 22, 18, 16, 35, 25]

for i, (header, width) in enumerate(zip(daily_headers, col_widths), 1):
    cell = ws_daily.cell(row=row, column=i, value=header)
    style_cell(cell, font=Font(name="Arial", bold=True, size=12, color="FFFFFF"), fill=title_fill)
    ws_daily.column_dimensions[get_column_letter(i)].width = width
ws_daily.row_dimensions[row].height = 30
header_row = row
row += 1

# Pre-fill 500 empty rows with formatting
data_start_row = row
NUM_ROWS = 500
for r in range(data_start_row, data_start_row + NUM_ROWS):
    # Date
    cell = ws_daily.cell(row=r, column=1)
    cell.number_format = 'DD/MM/YYYY'
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

    # Amount
    cell = ws_daily.cell(row=r, column=2)
    cell.number_format = currency_format
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

    # Category, payment, who paid, description, notes
    for c in range(3, 8):
        cell = ws_daily.cell(row=r, column=c)
        cell.font = normal_font
        cell.alignment = rtl_align if c >= 6 else center_align
        cell.border = thin_border

    # Alternate row coloring
    if r % 2 == 0:
        for c in range(1, 8):
            ws_daily.cell(row=r, column=c).fill = light_gray_fill
data_end_row = data_start_row + NUM_ROWS - 1

# ============================================================
# Hidden settings sheet for dropdown lists
# ============================================================
ws_settings = wb.create_sheet("הגדרות")
ws_settings.sheet_view.rightToLeft = True

# Categories in column A
ws_settings.cell(row=1, column=1, value="קטגוריות").font = Font(name="Arial", bold=True)
for i, cat in enumerate(CATEGORIES, start=2):
    ws_settings.cell(row=i, column=1, value=cat).font = normal_font
ws_settings.column_dimensions["A"].width = 30
cat_last_row = len(CATEGORIES) + 1

# Payment methods in column B
ws_settings.cell(row=1, column=2, value="אמצעי תשלום").font = Font(name="Arial", bold=True)
for i, pay in enumerate(PAYMENT_METHODS, start=2):
    ws_settings.cell(row=i, column=2, value=pay).font = normal_font
ws_settings.column_dimensions["B"].width = 20
pay_last_row = len(PAYMENT_METHODS) + 1

# Spouses in column C
ws_settings.cell(row=1, column=3, value="מי שילם").font = Font(name="Arial", bold=True)
for i, sp in enumerate(SPOUSES, start=2):
    ws_settings.cell(row=i, column=3, value=sp).font = normal_font
ws_settings.column_dimensions["C"].width = 18
spouse_last_row = len(SPOUSES) + 1

ws_settings.sheet_state = "hidden"

# --- Dropdowns ---
dv_cat = DataValidation(type="list", formula1=f"=הגדרות!$A$2:$A${cat_last_row}", allow_blank=True)
dv_cat.prompt = "בחר קטגוריה"
dv_cat.promptTitle = "קטגוריה"
ws_daily.add_data_validation(dv_cat)
dv_cat.add(f"C{data_start_row}:C{data_end_row}")

dv_pay = DataValidation(type="list", formula1=f"=הגדרות!$B$2:$B${pay_last_row}", allow_blank=True)
dv_pay.prompt = "בחר אמצעי תשלום"
ws_daily.add_data_validation(dv_pay)
dv_pay.add(f"D{data_start_row}:D{data_end_row}")

dv_spouse = DataValidation(type="list", formula1=f"=הגדרות!$C$2:$C${spouse_last_row}", allow_blank=True)
dv_spouse.prompt = "מי שילם?"
ws_daily.add_data_validation(dv_spouse)
dv_spouse.add(f"E{data_start_row}:E{data_end_row}")

# Freeze & filter
ws_daily.freeze_panes = f"A{data_start_row}"
ws_daily.auto_filter.ref = f"A{header_row}:G{data_end_row}"

# ============================================================
# SHEET 2: הוצאות קבועות (Fixed Monthly)
# ============================================================
ws_fixed = wb.create_sheet("הוצאות קבועות")
ws_fixed.sheet_view.rightToLeft = True

fixed_expenses = {
    "דיור": [
        "משכנתא / שכר דירה",
        "ארנונה",
        "ועד בית",
        "ביטוח דירה",
    ],
    "חשבונות": [
        "חשמל",
        "מים",
        "גז",
        "אינטרנט",
        "טלפון נייד",
        "טלוויזיה / סטרימינג",
    ],
    "ביטוחים ופנסיה": [
        "ביטוח חיים",
        "פנסיה",
        "קרן השתלמות",
        "ביטוח בריאות משלים",
        "ביטוח רכב",
    ],
    "חינוך": [
        "גן / מעון / צהרון",
        "בית ספר",
        "חוגים",
    ],
    "הלוואות ותשלומים": [
        "הלוואה בנקאית",
        "תשלומים לאשראי",
        "ליסינג רכב",
    ],
    "חיסכון": [
        "חיסכון חודשי",
        "חיסכון לילדים",
        "קרן חירום",
    ],
}

income_items = [
    "משכורת - בן/בת זוג 1",
    "משכורת - בן/בת זוג 2",
    "קצבת ילדים",
    "הכנסות מהשכרה",
    "הכנסות נוספות",
]

row = 1

# Title
ws_fixed.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
cell = ws_fixed.cell(row=row, column=1, value="הכנסות והוצאות קבועות חודשיות")
style_cell(cell, font=Font(name="Arial", bold=True, size=18, color="FFFFFF"), fill=title_fill)
ws_fixed.row_dimensions[row].height = 40
row += 1

# Month headers
ws_fixed.column_dimensions["A"].width = 28
cell = ws_fixed.cell(row=row, column=1, value="פריט")
style_cell(cell, font=sub_header_font, fill=header_gray)

for i, month in enumerate(MONTHS_HEB):
    col = i + 2
    ws_fixed.column_dimensions[get_column_letter(col)].width = 13
    cell = ws_fixed.cell(row=row, column=col, value=month)
    style_cell(cell, font=sub_header_font, fill=header_gray)

ws_fixed.column_dimensions["N"].width = 15
cell = ws_fixed.cell(row=row, column=14, value="סה\"כ שנתי")
style_cell(cell, font=sub_header_font, fill=header_gray)
ws_fixed.row_dimensions[row].height = 25
row += 1

# ---- INCOME ----
ws_fixed.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
cell = ws_fixed.cell(row=row, column=1, value="הכנסות")
style_cell(cell, font=category_font, fill=income_fill)
for c in range(2, 15):
    ws_fixed.cell(row=row, column=c).fill = income_fill
    ws_fixed.cell(row=row, column=c).border = thin_border
ws_fixed.row_dimensions[row].height = 28
row += 1

inc_rows = []
for idx, item in enumerate(income_items):
    cell = ws_fixed.cell(row=row, column=1, value=item)
    style_cell(cell, align=rtl_align)
    if idx % 2 == 1:
        cell.fill = light_gray_fill
    for c in range(2, 14):
        cell = ws_fixed.cell(row=row, column=c)
        style_cell(cell, fmt=currency_format)
        if idx % 2 == 1:
            cell.fill = light_gray_fill
    cell = ws_fixed.cell(row=row, column=14)
    cell.value = f"=SUM(B{row}:M{row})"
    style_cell(cell, font=Font(name="Arial", bold=True, size=11), fmt=currency_format)
    if idx % 2 == 1:
        cell.fill = light_gray_fill
    inc_rows.append(row)
    row += 1

# Income subtotal
cell = ws_fixed.cell(row=row, column=1, value="סה\"כ הכנסות")
style_cell(cell, font=total_font, fill=subtotal_fill, align=rtl_align)
for c in range(2, 15):
    parts = "+".join(f"{get_column_letter(c)}{r}" for r in inc_rows)
    cell = ws_fixed.cell(row=row, column=c)
    cell.value = f"={parts}"
    style_cell(cell, font=total_font, fill=subtotal_fill, fmt=currency_format)
income_total_row = row
ws_fixed.row_dimensions[row].height = 25
row += 2

# ---- FIXED EXPENSES ----
cat_colors = {
    "דיור": "C62828",
    "חשבונות": "E65100",
    "ביטוחים ופנסיה": "37474F",
    "חינוך": "1565C0",
    "הלוואות ותשלומים": "6A1B9A",
    "חיסכון": "1B5E20",
}

expense_subtotal_rows = []
for cat_name, items in fixed_expenses.items():
    color = cat_colors.get(cat_name, "546E7A")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws_fixed.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws_fixed.cell(row=row, column=1, value=cat_name)
    style_cell(cell, font=category_font, fill=fill, align=rtl_align)
    for c in range(2, 15):
        ws_fixed.cell(row=row, column=c).fill = fill
        ws_fixed.cell(row=row, column=c).border = thin_border
    ws_fixed.row_dimensions[row].height = 28
    row += 1

    item_rows = []
    for idx, item in enumerate(items):
        cell = ws_fixed.cell(row=row, column=1, value=item)
        style_cell(cell, align=rtl_align)
        if idx % 2 == 1:
            cell.fill = light_gray_fill
        for c in range(2, 14):
            cell = ws_fixed.cell(row=row, column=c)
            style_cell(cell, fmt=currency_format)
            if idx % 2 == 1:
                cell.fill = light_gray_fill
        cell = ws_fixed.cell(row=row, column=14)
        cell.value = f"=SUM(B{row}:M{row})"
        style_cell(cell, font=Font(name="Arial", bold=True, size=11), fmt=currency_format)
        if idx % 2 == 1:
            cell.fill = light_gray_fill
        item_rows.append(row)
        row += 1

    cell = ws_fixed.cell(row=row, column=1, value=f"סה\"כ {cat_name}")
    style_cell(cell, font=total_font, fill=subtotal_fill, align=rtl_align)
    for c in range(2, 15):
        parts = "+".join(f"{get_column_letter(c)}{r}" for r in item_rows)
        cell = ws_fixed.cell(row=row, column=c)
        cell.value = f"={parts}"
        style_cell(cell, font=total_font, fill=subtotal_fill, fmt=currency_format)
    expense_subtotal_rows.append(row)
    ws_fixed.row_dimensions[row].height = 25
    row += 2

# Total fixed expenses
cell = ws_fixed.cell(row=row, column=1, value="סה\"כ הוצאות קבועות")
style_cell(cell, font=grand_total_font, fill=dark_red_fill, align=rtl_align)
for c in range(2, 15):
    parts = "+".join(f"{get_column_letter(c)}{r}" for r in expense_subtotal_rows)
    cell = ws_fixed.cell(row=row, column=c)
    cell.value = f"={parts}"
    style_cell(cell, font=grand_total_font, fill=dark_red_fill, fmt=currency_format)
fixed_total_row = row
ws_fixed.row_dimensions[row].height = 30
row += 1

ws_fixed.freeze_panes = "B3"

# ============================================================
# SHEET 3: סיכום חודשי (Monthly Summary) + Charts
# ============================================================
ws_summary = wb.create_sheet("סיכום חודשי")
ws_summary.sheet_view.rightToLeft = True

row = 1
ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
cell = ws_summary.cell(row=row, column=1, value="סיכום חודשי - הכנסות מול הוצאות")
style_cell(cell, font=Font(name="Arial", bold=True, size=18, color="FFFFFF"), fill=title_fill)
ws_summary.row_dimensions[row].height = 40
row += 1

# Headers
ws_summary.column_dimensions["A"].width = 30
cell = ws_summary.cell(row=row, column=1, value="")
style_cell(cell, fill=header_gray)
for i, month in enumerate(MONTHS_HEB):
    col = i + 2
    ws_summary.column_dimensions[get_column_letter(col)].width = 14
    cell = ws_summary.cell(row=row, column=col, value=month)
    style_cell(cell, font=sub_header_font, fill=header_gray)
ws_summary.column_dimensions["N"].width = 15
cell = ws_summary.cell(row=row, column=14, value="סה\"כ שנתי")
style_cell(cell, font=sub_header_font, fill=header_gray)
ws_summary.row_dimensions[row].height = 25
row += 1

# Total Income
cell = ws_summary.cell(row=row, column=1, value="סה\"כ הכנסות")
style_cell(cell, font=Font(name="Arial", bold=True, size=12, color="2E7D32"), fill=green_fill, align=rtl_align)
for c in range(2, 15):
    cell = ws_summary.cell(row=row, column=c)
    cell.value = f"='הוצאות קבועות'!{get_column_letter(c)}{income_total_row}"
    style_cell(cell, font=Font(name="Arial", bold=True, size=12, color="2E7D32"), fill=green_fill, fmt=currency_format)
income_summary_row = row
ws_summary.row_dimensions[row].height = 28
row += 1

# Fixed expenses
cell = ws_summary.cell(row=row, column=1, value="הוצאות קבועות")
style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="B71C1C"), fill=red_fill_light, align=rtl_align)
for c in range(2, 15):
    cell = ws_summary.cell(row=row, column=c)
    cell.value = f"='הוצאות קבועות'!{get_column_letter(c)}{fixed_total_row}"
    style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="B71C1C"), fill=red_fill_light, fmt=currency_format)
fixed_summary_row = row
ws_summary.row_dimensions[row].height = 28
row += 1

# Daily expenses per month
cell = ws_summary.cell(row=row, column=1, value="הוצאות יומיות (מהיומן)")
style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="E65100"), fill=orange_fill_light, align=rtl_align)
for c in range(2, 14):
    month_num = c - 1
    cell = ws_summary.cell(row=row, column=c)
    cell.value = (f"=SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})"
                  f"*(MONTH('יומן הוצאות'!A{data_start_row}:A{data_end_row})={month_num}))")
    style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="E65100"), fill=orange_fill_light, fmt=currency_format)
cell = ws_summary.cell(row=row, column=14)
cell.value = f"=SUM(B{row}:M{row})"
style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="E65100"), fill=orange_fill_light, fmt=currency_format)
daily_summary_row = row
ws_summary.row_dimensions[row].height = 28
row += 1

# Total all expenses
cell = ws_summary.cell(row=row, column=1, value="סה\"כ כל ההוצאות")
style_cell(cell, font=grand_total_font, fill=dark_red_fill, align=rtl_align)
for c in range(2, 15):
    cl = get_column_letter(c)
    cell = ws_summary.cell(row=row, column=c)
    cell.value = f"={cl}{fixed_summary_row}+{cl}{daily_summary_row}"
    style_cell(cell, font=grand_total_font, fill=dark_red_fill, fmt=currency_format)
total_all_expenses_row = row
ws_summary.row_dimensions[row].height = 30
row += 2

# Balance — with conditional formatting (1)
cell = ws_summary.cell(row=row, column=1, value="יתרה חודשית (הכנסות - הוצאות)")
style_cell(cell, font=grand_total_font, fill=summary_fill, align=rtl_align)
for c in range(2, 15):
    cl = get_column_letter(c)
    cell = ws_summary.cell(row=row, column=c)
    cell.value = f"={cl}{income_summary_row}-{cl}{total_all_expenses_row}"
    style_cell(cell, font=grand_total_font, fill=summary_fill, fmt=currency_format)
balance_row = row
ws_summary.row_dimensions[row].height = 32

# (1) Conditional formatting: green if positive, red if negative
green_font_cf = Font(name="Arial", bold=True, size=13, color="1B5E20")
red_font_cf = Font(name="Arial", bold=True, size=13, color="B71C1C")
balance_range = f"B{balance_row}:N{balance_row}"
ws_summary.conditional_formatting.add(
    balance_range,
    CellIsRule(operator="greaterThan", formula=["0"],
              fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
              font=green_font_cf)
)
ws_summary.conditional_formatting.add(
    balance_range,
    CellIsRule(operator="lessThan", formula=["0"],
              fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
              font=red_font_cf)
)
row += 1

# Savings rate
cell = ws_summary.cell(row=row, column=1, value="אחוז חיסכון")
style_cell(cell, font=Font(name="Arial", bold=True, size=12, color="1B5E20"), fill=green_fill, align=rtl_align)
for c in range(2, 15):
    cl = get_column_letter(c)
    cell = ws_summary.cell(row=row, column=c)
    cell.value = f'=IF({cl}{income_summary_row}=0,"",{cl}{balance_row}/{cl}{income_summary_row})'
    style_cell(cell, font=Font(name="Arial", bold=True, size=12, color="1B5E20"), fill=green_fill, fmt='0.0%')
savings_rate_row = row
# Conditional formatting on savings rate too
ws_summary.conditional_formatting.add(
    f"B{savings_rate_row}:N{savings_rate_row}",
    CellIsRule(operator="greaterThan", formula=["0"],
              fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
              font=Font(name="Arial", bold=True, size=12, color="1B5E20"))
)
ws_summary.conditional_formatting.add(
    f"B{savings_rate_row}:N{savings_rate_row}",
    CellIsRule(operator="lessThan", formula=["0"],
              fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
              font=Font(name="Arial", bold=True, size=12, color="B71C1C"))
)
ws_summary.row_dimensions[row].height = 28
row += 2

# (2) סיכום יומי — "כמה הוצאתי היום"
cell = ws_summary.cell(row=row, column=1, value="סה\"כ הוצאות היום")
style_cell(cell, font=Font(name="Arial", bold=True, size=14, color="FFFFFF"),
           fill=PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid"), align=rtl_align)
ws_summary.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
cell = ws_summary.cell(row=row, column=2)
cell.value = (f"=SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})"
              f"*('יומן הוצאות'!A{data_start_row}:A{data_end_row}=TODAY()))")
style_cell(cell, font=Font(name="Arial", bold=True, size=16, color="FFFFFF"),
           fill=PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid"), fmt=currency_format)
ws_summary.cell(row=row, column=3).fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
today_row = row
ws_summary.row_dimensions[row].height = 35
row += 2

# ---- Breakdown: daily expenses by category per month ----
cell = ws_summary.cell(row=row, column=1, value="פירוט הוצאות יומיות לפי קטגוריה")
style_cell(cell, font=Font(name="Arial", bold=True, size=14, color="FFFFFF"), fill=title_fill)
ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
for c in range(2, 15):
    ws_summary.cell(row=row, column=c).fill = title_fill
    ws_summary.cell(row=row, column=c).border = thin_border
ws_summary.row_dimensions[row].height = 32
row += 1

# Category headers
cell = ws_summary.cell(row=row, column=1, value="קטגוריה")
style_cell(cell, font=sub_header_font, fill=header_gray)
for i, month in enumerate(MONTHS_HEB):
    cell = ws_summary.cell(row=row, column=i + 2, value=month)
    style_cell(cell, font=sub_header_font, fill=header_gray)
cell = ws_summary.cell(row=row, column=14, value="סה\"כ שנתי")
style_cell(cell, font=sub_header_font, fill=header_gray)
cat_header_row = row
row += 1

cat_data_start_row = row
for idx, cat in enumerate(CATEGORIES):
    cell = ws_summary.cell(row=row, column=1, value=cat)
    style_cell(cell, align=rtl_align)
    if idx % 2 == 1:
        cell.fill = light_gray_fill

    for c in range(2, 14):
        month_num = c - 1
        cell = ws_summary.cell(row=row, column=c)
        cell.value = (
            f"=SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})"
            f"*(MONTH('יומן הוצאות'!A{data_start_row}:A{data_end_row})={month_num})"
            f"*('יומן הוצאות'!C{data_start_row}:C{data_end_row}=\"{cat}\"))"
        )
        style_cell(cell, fmt=currency_format)
        if idx % 2 == 1:
            cell.fill = light_gray_fill

    cell = ws_summary.cell(row=row, column=14)
    cell.value = f"=SUM(B{row}:M{row})"
    style_cell(cell, font=Font(name="Arial", bold=True, size=11), fmt=currency_format)
    if idx % 2 == 1:
        cell.fill = light_gray_fill
    row += 1
cat_data_end_row = row - 1

# (7) Breakdown by spouse
row += 1
cell = ws_summary.cell(row=row, column=1, value="פירוט הוצאות יומיות לפי מי שילם")
style_cell(cell, font=Font(name="Arial", bold=True, size=14, color="FFFFFF"), fill=title_fill)
ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
for c in range(2, 15):
    ws_summary.cell(row=row, column=c).fill = title_fill
    ws_summary.cell(row=row, column=c).border = thin_border
ws_summary.row_dimensions[row].height = 32
row += 1

cell = ws_summary.cell(row=row, column=1, value="מי שילם")
style_cell(cell, font=sub_header_font, fill=header_gray)
for i, month in enumerate(MONTHS_HEB):
    cell = ws_summary.cell(row=row, column=i + 2, value=month)
    style_cell(cell, font=sub_header_font, fill=header_gray)
cell = ws_summary.cell(row=row, column=14, value="סה\"כ שנתי")
style_cell(cell, font=sub_header_font, fill=header_gray)
row += 1

spouse_data_start_row = row
for idx, sp in enumerate(SPOUSES):
    cell = ws_summary.cell(row=row, column=1, value=sp)
    style_cell(cell, align=rtl_align)
    if idx % 2 == 1:
        cell.fill = light_gray_fill

    for c in range(2, 14):
        month_num = c - 1
        cell = ws_summary.cell(row=row, column=c)
        cell.value = (
            f"=SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})"
            f"*(MONTH('יומן הוצאות'!A{data_start_row}:A{data_end_row})={month_num})"
            f"*('יומן הוצאות'!E{data_start_row}:E{data_end_row}=\"{sp}\"))"
        )
        style_cell(cell, fmt=currency_format)
        if idx % 2 == 1:
            cell.fill = light_gray_fill

    cell = ws_summary.cell(row=row, column=14)
    cell.value = f"=SUM(B{row}:M{row})"
    style_cell(cell, font=Font(name="Arial", bold=True, size=11), fmt=currency_format)
    if idx % 2 == 1:
        cell.fill = light_gray_fill
    row += 1
spouse_data_end_row = row - 1

# (3) Charts
row += 2

# --- Pie Chart: yearly expenses by category ---
pie = PieChart()
pie.title = "התפלגות הוצאות לפי קטגוריה (שנתי)"
pie.style = 10
pie.width = 22
pie.height = 14

cat_labels = Reference(ws_summary, min_col=1, min_row=cat_data_start_row, max_row=cat_data_end_row)
cat_values = Reference(ws_summary, min_col=14, min_row=cat_data_start_row, max_row=cat_data_end_row)
pie.add_data(cat_values, titles_from_data=False)
pie.set_categories(cat_labels)
pie.dataLabels = openpyxl.chart.label.DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.dataLabels.showVal = False

ws_summary.add_chart(pie, f"A{row}")
pie_chart_row = row
row += 18

# --- Bar Chart: monthly income vs expenses ---
bar = BarChart()
bar.type = "col"
bar.title = "הכנסות מול הוצאות - לפי חודש"
bar.style = 10
bar.width = 28
bar.height = 14
bar.y_axis.title = "סכום (₪)"

# Income series
income_data = Reference(ws_summary, min_col=2, max_col=13, min_row=income_summary_row)
bar.add_data(income_data, from_rows=True, titles_from_data=False)
bar.series[0].title = openpyxl.chart.series.SeriesLabel(v="הכנסות")
bar.series[0].graphicalProperties.solidFill = "4CAF50"

# Total expenses series
expense_data = Reference(ws_summary, min_col=2, max_col=13, min_row=total_all_expenses_row)
bar.add_data(expense_data, from_rows=True, titles_from_data=False)
bar.series[1].title = openpyxl.chart.series.SeriesLabel(v="הוצאות")
bar.series[1].graphicalProperties.solidFill = "E53935"

# Balance series
balance_data = Reference(ws_summary, min_col=2, max_col=13, min_row=balance_row)
bar.add_data(balance_data, from_rows=True, titles_from_data=False)
bar.series[2].title = openpyxl.chart.series.SeriesLabel(v="יתרה")
bar.series[2].graphicalProperties.solidFill = "1565C0"

month_cats = Reference(ws_summary, min_col=2, max_col=13, min_row=2)
bar.set_categories(month_cats)

ws_summary.add_chart(bar, f"A{row}")
row += 18

# --- Bar Chart: by spouse ---
bar_spouse = BarChart()
bar_spouse.type = "col"
bar_spouse.title = "הוצאות לפי מי שילם - חודשי"
bar_spouse.style = 10
bar_spouse.width = 28
bar_spouse.height = 14
bar_spouse.y_axis.title = "סכום (₪)"

spouse_colors = ["1565C0", "E91E63", "FF9800"]
for idx, sp_row in enumerate(range(spouse_data_start_row, spouse_data_end_row + 1)):
    sp_data = Reference(ws_summary, min_col=2, max_col=13, min_row=sp_row)
    bar_spouse.add_data(sp_data, from_rows=True, titles_from_data=False)
    bar_spouse.series[idx].title = openpyxl.chart.series.SeriesLabel(v=SPOUSES[idx])
    bar_spouse.series[idx].graphicalProperties.solidFill = spouse_colors[idx]

bar_spouse.set_categories(month_cats)
ws_summary.add_chart(bar_spouse, f"A{row}")

ws_summary.freeze_panes = "B3"

# ============================================================
# SHEET 4: דשבורד (Dashboard) (6)
# ============================================================
ws_dash = wb.create_sheet("דשבורד")
ws_dash.sheet_view.rightToLeft = True

for c in range(1, 10):
    ws_dash.column_dimensions[get_column_letter(c)].width = 18

row = 1
ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
cell = ws_dash.cell(row=row, column=1, value="דשבורד - מבט מהיר")
style_cell(cell, font=Font(name="Arial", bold=True, size=22, color="FFFFFF"), fill=dashboard_fill)
for c in range(2, 9):
    ws_dash.cell(row=row, column=c).fill = dashboard_fill
ws_dash.row_dimensions[row].height = 50
row += 2

# --- Card style helper ---
def make_card(sheet, r, c, title, formula, fmt=currency_format, span=2):
    sheet.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    cell = sheet.cell(row=r, column=c, value=title)
    style_cell(cell, font=Font(name="Arial", bold=True, size=11, color="B0BEC5"), fill=card_fill, align=center_align)
    for cc in range(c + 1, c + span):
        sheet.cell(row=r, column=cc).fill = card_fill
        sheet.cell(row=r, column=cc).border = thin_border
    sheet.row_dimensions[r].height = 25

    sheet.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + span - 1)
    cell = sheet.cell(row=r + 1, column=c, value=formula)
    style_cell(cell, font=Font(name="Arial", bold=True, size=20, color="FFFFFF"), fill=card_fill, align=center_align, fmt=fmt)
    for cc in range(c + 1, c + span):
        sheet.cell(row=r + 1, column=cc).fill = card_fill
        sheet.cell(row=r + 1, column=cc).border = thin_border
    sheet.row_dimensions[r + 1].height = 45

# Current month number formula part
cm = "MONTH(TODAY())"
sm_sheet = "'סיכום חודשי'"

# Card 1: הוצאות היום
make_card(ws_dash, row, 1, "הוצאות היום",
          f"=SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})*('יומן הוצאות'!A{data_start_row}:A{data_end_row}=TODAY()))")

# Card 2: הוצאות החודש (daily + fixed)
make_card(ws_dash, row, 4, "סה\"כ הוצאות החודש",
          f"=INDEX({sm_sheet}!B{total_all_expenses_row}:M{total_all_expenses_row},1,{cm})")

# Card 3: הכנסות החודש
make_card(ws_dash, row, 7, "הכנסות החודש",
          f"=INDEX({sm_sheet}!B{income_summary_row}:M{income_summary_row},1,{cm})")
row += 3

# Card 4: יתרה החודש
make_card(ws_dash, row, 1, "יתרה החודש",
          f"=INDEX({sm_sheet}!B{balance_row}:M{balance_row},1,{cm})")
# Conditional formatting on balance card value
balance_card_cell = f"A{row+1}:B{row+1}"
ws_dash.conditional_formatting.add(
    balance_card_cell,
    CellIsRule(operator="greaterThan", formula=["0"],
              font=Font(name="Arial", bold=True, size=20, color="4CAF50"))
)
ws_dash.conditional_formatting.add(
    balance_card_cell,
    CellIsRule(operator="lessThan", formula=["0"],
              font=Font(name="Arial", bold=True, size=20, color="FF5252"))
)

# Card 5: ממוצע יומי החודש
make_card(ws_dash, row, 4, "ממוצע הוצאה יומית (החודש)",
          f"=IFERROR(SUMPRODUCT(('יומן הוצאות'!B{data_start_row}:B{data_end_row})*(MONTH('יומן הוצאות'!A{data_start_row}:A{data_end_row})={cm}))/DAY(TODAY()),0)")

# Card 6: אחוז חיסכון
make_card(ws_dash, row, 7, "אחוז חיסכון החודש",
          f"=IFERROR(INDEX({sm_sheet}!B{balance_row}:M{balance_row},1,{cm})/INDEX({sm_sheet}!B{income_summary_row}:M{income_summary_row},1,{cm}),0)",
          fmt='0.0%')
row += 3

# Card 7: הקטגוריה הכי יקרה החודש
make_card(ws_dash, row, 1, "הקטגוריה הכי יקרה החודש",
          f'=IFERROR(INDEX({sm_sheet}!A{cat_data_start_row}:A{cat_data_end_row},MATCH(MAX(INDEX({sm_sheet}!B{cat_data_start_row}:M{cat_data_end_row},,{cm})),INDEX({sm_sheet}!B{cat_data_start_row}:M{cat_data_end_row},,{cm}),0)),"")',
          fmt='@', span=3)

# Card 8: סכום הקטגוריה הכי יקרה
make_card(ws_dash, row, 5, "סכום הקטגוריה הכי יקרה",
          f"=IFERROR(MAX(INDEX({sm_sheet}!B{cat_data_start_row}:M{cat_data_end_row},,{cm})),0)",
          span=3)
row += 3

# Card 9: סה"כ שנתי
make_card(ws_dash, row, 1, "סה\"כ הוצאות שנתי",
          f"={sm_sheet}!N{total_all_expenses_row}")
make_card(ws_dash, row, 4, "סה\"כ הכנסות שנתי",
          f"={sm_sheet}!N{income_summary_row}")
make_card(ws_dash, row, 7, "יתרה שנתית",
          f"={sm_sheet}!N{balance_row}")

# ============================================================
# Sheet order
# ============================================================
wb.move_sheet("דשבורד", offset=-3)  # make dashboard first

# ============================================================
# Save
# ============================================================
output_path = r"c:\Users\amith\ads-dashboard\תקציב_משק_בית_2026.xlsx"
wb.save(output_path)
print("File saved successfully!")
