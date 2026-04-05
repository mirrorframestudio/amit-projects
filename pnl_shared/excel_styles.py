"""
pnl_shared.excel_styles — shared Excel style constants and helper functions
used by both monday-pnl and weekly-pnl.

Both projects previously defined identical (or near-identical) fonts, fills,
borders, and RTL helpers. This module is the single source of truth.

Usage:
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
    from pnl_shared.excel_styles import (
        HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
        THIN_BORDER, PROFIT_POS_FONT, PROFIT_NEG_FONT,
        CURRENCY_FMT, PERCENT_FMT,
        apply_header_row, set_rtl, auto_width, write_kpi_row,
    )
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Brand colours ────────────────────────────────────────────
_BLUE_DARK = "2F5496"      # header background
_BLUE_MID = "D6E4F0"       # KPI label background
_BLUE_NAVY = "1F3864"      # month-group header background
_GREEN_CELL = "E2EFDA"     # subtotal / positive highlight fill
_GREEN_TEXT = "006100"     # positive profit text
_RED_TEXT = "9C0006"       # negative profit text
_RED_CELL = "FFC7CE"       # warning / negative fill

# ── Fonts ────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
KPI_FONT = Font(name="Calibri", bold=True, size=12)
SECTION_FONT = Font(name="Calibri", bold=True, size=13, color=_BLUE_DARK)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=_BLUE_DARK)

PROFIT_POS_FONT = Font(name="Calibri", size=10, color=_GREEN_TEXT)
PROFIT_NEG_FONT = Font(name="Calibri", size=10, color=_RED_TEXT)

MONTH_HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
MONTH_SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)

# ── Fills ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color=_BLUE_DARK, end_color=_BLUE_DARK, fill_type="solid")
KPI_LABEL_FILL = PatternFill(start_color=_BLUE_MID, end_color=_BLUE_MID, fill_type="solid")
MONTH_HEADER_FILL = PatternFill(start_color=_BLUE_NAVY, end_color=_BLUE_NAVY, fill_type="solid")
MONTH_SUBTOTAL_FILL = PatternFill(start_color=_GREEN_CELL, end_color=_GREEN_CELL, fill_type="solid")
WARNING_FILL = PatternFill(start_color=_RED_CELL, end_color=_RED_CELL, fill_type="solid")

# ── Alignments ───────────────────────────────────────────────
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# ── Borders ──────────────────────────────────────────────────
_THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

# ── Number formats ───────────────────────────────────────────
CURRENCY_FMT = '₪#,##0.00'
CURRENCY_FMT_COMPACT = '₪#,##0'
PERCENT_FMT = '0.0%'
PERCENT_FMT_INT = '0%'

# ── Hebrew month names ───────────────────────────────────────
HEBREW_MONTHS = {
    "01": "ינואר", "02": "פברואר", "03": "מרץ",
    "04": "אפריל", "05": "מאי",    "06": "יוני",
    "07": "יולי",  "08": "אוגוסט", "09": "ספטמבר",
    "10": "אוקטובר", "11": "נובמבר", "12": "דצמבר",
}


def month_heb(month_key: str) -> str:
    """Convert 'YYYY-MM' to Hebrew month label, e.g. '2026-03' → 'מרץ 2026'."""
    try:
        year, month = month_key.split("-")
        return f"{HEBREW_MONTHS.get(month, month)} {year}"
    except (ValueError, AttributeError):
        return month_key


# ── Worksheet helpers ────────────────────────────────────────

def set_rtl(ws: Worksheet) -> None:
    """Set the worksheet to right-to-left (Hebrew) direction."""
    ws.sheet_view.rightToLeft = True


def auto_width(ws: Worksheet, min_width: int = 10, max_width: int = 40) -> None:
    """Auto-fit all column widths based on cell content length."""
    for col_cells in ws.columns:
        lengths = [len(str(cell.value)) for cell in col_cells if cell.value is not None]
        best = max(lengths) + 2 if lengths else min_width
        best = max(min_width, min(best, max_width))
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = best


def apply_header_row(
    ws: Worksheet,
    headers: list[str],
    row: int = 1,
    start_col: int = 1,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
    border: Border | None = None,
    height: float = 20,
) -> None:
    """
    Write a styled header row.

    Args:
        ws:         Target worksheet.
        headers:    List of header label strings.
        row:        Row number (1-based). Default: 1.
        start_col:  Starting column (1-based). Default: 1.
        font:       Override font (default: HEADER_FONT).
        fill:       Override fill (default: HEADER_FILL).
        align:      Override alignment (default: HEADER_ALIGN).
        border:     Override border (default: THIN_BORDER).
        height:     Row height in points. Default: 20.
    """
    _font = font or HEADER_FONT
    _fill = fill or HEADER_FILL
    _align = align or HEADER_ALIGN
    _border = border or THIN_BORDER

    ws.row_dimensions[row].height = height
    for col_offset, label in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + col_offset, value=label)
        cell.font = _font
        cell.fill = _fill
        cell.alignment = _align
        cell.border = _border


def write_kpi_row(
    ws: Worksheet,
    label: str,
    value,
    row: int,
    label_col: int = 1,
    value_col: int = 2,
    number_format: str | None = None,
) -> None:
    """
    Write a single KPI label + value pair with standard styling.

    Args:
        ws:            Target worksheet.
        label:         KPI label text.
        value:         KPI value (number, string, etc.).
        row:           Row number to write into.
        label_col:     Column for the label. Default: 1.
        value_col:     Column for the value. Default: 2.
        number_format: Optional number format string for the value cell.
    """
    label_cell = ws.cell(row=row, column=label_col, value=label)
    label_cell.font = KPI_FONT
    label_cell.fill = KPI_LABEL_FILL
    label_cell.border = THIN_BORDER
    label_cell.alignment = RIGHT_ALIGN

    value_cell = ws.cell(row=row, column=value_col, value=value)
    value_cell.font = BODY_FONT
    value_cell.border = THIN_BORDER
    value_cell.alignment = CENTER_ALIGN
    if number_format:
        value_cell.number_format = number_format


def apply_profit_color(cell, value: float) -> None:
    """Color a profit cell green (positive) or red (negative)."""
    if value >= 0:
        cell.font = PROFIT_POS_FONT
    else:
        cell.font = PROFIT_NEG_FONT
