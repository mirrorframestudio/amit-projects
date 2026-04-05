"""
Excel writer — creates/updates P&L.xlsx with 4 sheets:
  1. הגדרות     — all parameters
  2. הזמנות_גולמי — raw Monday data
  3. חישוב       — orders + computed P&L columns
  4. דשבורד      — KPIs + charts + monthly/status/shipping/product summaries

Supports upsert: existing orders (matched by order_id) are updated in place;
new orders are appended.
"""

import logging
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import settings

# ── Shared styles (single source of truth for all P&L Excel files) ──
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from pnl_shared.excel_styles import (  # noqa: E402
    HEADER_FONT as _HEADER_FONT,
    HEADER_FILL as _HEADER_FILL,
    HEADER_ALIGN as _HEADER_ALIGN,
    THIN_BORDER as _THIN_BORDER,
    BODY_FONT as _BODY_FONT,
    KPI_FONT as _KPI_FONT,
    KPI_LABEL_FILL as _KPI_LABEL_FILL,
    SECTION_FONT as _SECTION_FONT,
    PROFIT_POS_FONT as _PROFIT_POS_FONT,
    PROFIT_NEG_FONT as _PROFIT_NEG_FONT,
    MONTH_HEADER_FONT as _MONTH_HEADER_FONT,
    MONTH_HEADER_FILL as _MONTH_HEADER_FILL,
    MONTH_SUBTOTAL_FONT as _MONTH_SUBTOTAL_FONT,
    MONTH_SUBTOTAL_FILL as _MONTH_SUBTOTAL_FILL,
    HEBREW_MONTHS as _HEBREW_MONTHS,
    CURRENCY_FMT,
    PERCENT_FMT,
    set_rtl,
    auto_width,
    apply_header_row as _apply_header_row_shared,
)

logger = logging.getLogger(__name__)

# ── Column definitions (matched to your board) ─────────────────
RAW_COLUMNS = [
    "order_id", "customer_name", "order_date", "status", "order_price",
    "shipping_type", "product_type", "phone", "email", "location",
    "tracking_num", "deadline", "notes",
    "num_products", "total_qty", "product_names", "created_at", "last_synced_at",
]

RAW_COLUMNS_HEB = [
    "מס׳ הזמנה", "שם לקוח", "תאריך הזמנה", "סטטוס", "מחיר הזמנה",
    "סוג משלוח", "סוג מוצר", "טלפון", "אימייל", "כתובת",
    "מס׳ מעקב", "דדליין", "הערות",
    "מס׳ מוצרים", "כמות כוללת", "שמות מוצרים", "תאריך יצירה", "עדכון אחרון",
]

CALC_COLUMNS = RAW_COLUMNS + [
    "sell_price", "vat", "cogs", "shipping", "payment_fee", "profit", "profit_margin",
]

CALC_COLUMNS_HEB = RAW_COLUMNS_HEB + [
    "מחיר מכירה", "מע״מ", "עלות סחורה", "משלוח", "עמלת סליקה", "רווח", "אחוז רווח",
]

CURRENCY_COLS = {"order_price", "sell_price", "vat", "cogs", "shipping", "payment_fee", "profit"}

PERCENT_COLS = {"profit_margin"}

def _month_heb(month_key: str) -> str:
    """Convert YYYY-MM to Hebrew month name + year, e.g. 'ינואר 2026'."""
    if month_key == "Unknown" or len(month_key) < 7:
        return "לא ידוע"
    mm = month_key[5:7]
    yyyy = month_key[:4]
    return f"{_HEBREW_MONTHS.get(mm, mm)} {yyyy}"


def _apply_header_style(ws: Worksheet, columns: list[str], heb_columns: list[str] | None = None) -> None:
    display = heb_columns or columns
    for col_idx, col_name in enumerate(display, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(display))}1"


def _set_column_widths(ws: Worksheet, columns: list[str]) -> None:
    for col_idx, col_name in enumerate(columns, 1):
        width = max(12, min(24, len(col_name) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _currency_fmt() -> str:
    return '"₪"#,##0.00' if settings.CURRENCY == "ILS" else '"$"#,##0.00'


def _write_data_row(
    ws: Worksheet, row_num: int, order: dict[str, Any], columns: list[str],
) -> None:
    fmt = _currency_fmt()
    for col_idx, col_name in enumerate(columns, 1):
        value = order.get(col_name)
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER

        if col_name in CURRENCY_COLS and isinstance(value, (int, float)):
            cell.number_format = fmt
        elif col_name in PERCENT_COLS and isinstance(value, (int, float)):
            cell.number_format = "0.00%"
        elif col_name == "order_date":
            cell.number_format = "YYYY-MM-DD"

        if col_name == "profit" and isinstance(value, (int, float)):
            cell.font = _PROFIT_POS_FONT if value >= 0 else _PROFIT_NEG_FONT


# ── Sheet builders ──────────────────────────────────────────────

def _build_settings_sheet(wb: Workbook) -> None:
    if "הגדרות" in wb.sheetnames:
        del wb["הגדרות"]
    ws = wb.create_sheet("הגדרות", 0)
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:C1")
    ws["A1"].value = "הגדרות רווח והפסד — OneZoneJersey"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="עדכון אחרון").font = Font(italic=True, size=9)
    ws.cell(row=2, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))

    for col_idx, header in enumerate(["פרמטר", "ערך", "תיאור"], 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER

    params = [
        ("מזהה בורד Monday", settings.MONDAY_BOARD_ID, "מזהה הבורד ב-Monday.com"),
        ("עלות חולצה רגילה", settings.COST_SHIRT, "עלות בסיס: חולצה רגילה ($)"),
        ("עלות חולצה רטרו", settings.COST_RETRO, "עלות בסיס: חולצה רטרו ($)"),
        ("עלות גרסת שחקן", settings.COST_PLAYER_VERSION, "עלות בסיס: גרסת שחקן ($)"),
        ("עלות שם ומספר", settings.COST_NAME_NUMBER, "תוספת: שם ומספר ($)"),
        ("עלות מכנסיים", settings.COST_PANTS, "תוספת: מכנסיים ($)"),
        ("עלות שרוול ארוך", settings.COST_LONG_SHIRT, "תוספת: שרוול ארוך ($)"),
        ("עלות גרביים", settings.COST_SOCKS, "תוספת: גרביים ($)"),
        ("עלות פאצ׳", settings.COST_PATCH, "תוספת: פאצ׳ ($)"),
        ("עלות משלוח לבית", settings.COST_HOME_DELIVERY, "משלוח: עד הבית לכל הזמנה ($)"),
        ("שיעור מע״מ", settings.VAT_RATE, "שיעור מע״מ (0.18 = 18%)"),
        ("עמלת סליקה", settings.PAYMENT_FEE_RATE, "אחוז עמלת סליקה"),
        ("מצב חישוב", settings.MODE, "INCLUDES_VAT = מחיר כולל מע״מ"),
        ("מטבע", settings.CURRENCY, "מטבע (ILS או USD)"),
    ]

    for row_idx, (param, value, desc) in enumerate(params, 5):
        ws.cell(row=row_idx, column=1, value=param).font = Font(bold=True, size=10)
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.font = _BODY_FONT
        val_cell.border = _THIN_BORDER
        if isinstance(value, float) and value < 1:
            val_cell.number_format = "0.00%"
        ws.cell(row=row_idx, column=3, value=desc).font = Font(italic=True, size=9, color="666666")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40


def _get_or_create_sheet(
    wb: Workbook, name: str, columns: list[str], heb_columns: list[str] | None = None,
) -> tuple[Worksheet, dict[str, int]]:
    if name in wb.sheetnames:
        ws = wb[name]
        # Refresh header row in case columns were added/reordered
        _apply_header_style(ws, columns, heb_columns)
        _set_column_widths(ws, columns)
        existing_ids: dict[str, int] = {}
        oid_col = columns.index("order_id") + 1
        for row_num in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=oid_col).value
            if cell_val is not None:
                existing_ids[str(cell_val)] = row_num
        return ws, existing_ids
    else:
        ws = wb.create_sheet(name)
        ws.sheet_view.rightToLeft = True
        _apply_header_style(ws, columns, heb_columns)
        _set_column_widths(ws, columns)
        return ws, {}


def _upsert_rows(
    ws: Worksheet, orders: list[dict[str, Any]], columns: list[str],
    existing_ids: dict[str, int],
) -> None:
    next_row = ws.max_row + 1 if ws.max_row > 1 else 2
    updated = 0
    inserted = 0

    for order in orders:
        oid = str(order.get("order_id", ""))
        if oid in existing_ids:
            _write_data_row(ws, existing_ids[oid], order, columns)
            updated += 1
        else:
            _write_data_row(ws, next_row, order, columns)
            existing_ids[oid] = next_row
            next_row += 1
            inserted += 1

    logger.info("Sheet '%s': %d updated, %d inserted", ws.title, updated, inserted)


def _build_raw_sheet(wb: Workbook, orders: list[dict[str, Any]]) -> None:
    ws, existing_ids = _get_or_create_sheet(wb, "הזמנות_גולמי", RAW_COLUMNS, RAW_COLUMNS_HEB)
    _upsert_rows(ws, orders, RAW_COLUMNS, existing_ids)


def _get_month_key(order: dict[str, Any]) -> str:
    """Extract YYYY-MM from order_date."""
    d = order.get("order_date")
    if d and hasattr(d, "strftime"):
        return d.strftime("%Y-%m")
    elif d:
        s = str(d)[:7]
        return s if len(s) == 7 else "Unknown"
    return "Unknown"


def _build_calc_sheet(wb: Workbook, orders: list[dict[str, Any]]) -> None:
    """Build Calc sheet with orders grouped by month, with headers and subtotals."""
    if "חישוב" in wb.sheetnames:
        del wb["חישוב"]
    ws = wb.create_sheet("חישוב")
    ws.sheet_view.rightToLeft = True

    num_cols = len(CALC_COLUMNS)
    fmt = _currency_fmt()

    # Column headers in row 1 (Hebrew)
    _apply_header_style(ws, CALC_COLUMNS, CALC_COLUMNS_HEB)
    _set_column_widths(ws, CALC_COLUMNS)

    # Group orders by month
    monthly_orders: dict[str, list[dict]] = OrderedDict()
    for o in orders:
        mk = _get_month_key(o)
        monthly_orders.setdefault(mk, []).append(o)

    # Sort months descending (newest first)
    sorted_months = sorted(monthly_orders.keys(), reverse=True)

    row = 2
    for month_key in sorted_months:
        month_orders = monthly_orders[month_key]

        # ── Month header row ────────────────────────────────────
        m_profit = sum(o.get("profit", 0) or 0 for o in month_orders)
        m_count = len(month_orders)

        header_text = f"{_month_heb(month_key)}  |  {m_count} הזמנות  |  רווח: ₪{m_profit:,.0f}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        hcell = ws.cell(row=row, column=1, value=header_text)
        hcell.font = _MONTH_HEADER_FONT
        hcell.fill = _MONTH_HEADER_FILL
        hcell.alignment = Alignment(horizontal="right", vertical="center")
        for ci in range(2, num_cols + 1):
            c = ws.cell(row=row, column=ci)
            c.fill = _MONTH_HEADER_FILL
        row += 1

        # ── Order rows ──────────────────────────────────────────
        for o in month_orders:
            _write_data_row(ws, row, o, CALC_COLUMNS)
            row += 1

        # ── Subtotal row ────────────────────────────────────────
        m_price = sum(o.get("order_price", 0) or 0 for o in month_orders)
        m_vat = sum(o.get("vat", 0) or 0 for o in month_orders)
        m_cogs = sum(o.get("cogs", 0) or 0 for o in month_orders)
        m_ship = sum(o.get("shipping", 0) or 0 for o in month_orders)
        m_fee = sum(o.get("payment_fee", 0) or 0 for o in month_orders)
        m_margin = m_profit / m_price if m_price else 0

        subtotals = {
            "order_id": f"סה״כ {_month_heb(month_key)}",
            "order_price": m_price,
            "sell_price": m_price,
            "vat": m_vat,
            "cogs": m_cogs,
            "shipping": m_ship,
            "payment_fee": m_fee,
            "profit": m_profit,
            "profit_margin": m_margin,
            "num_products": sum(o.get("num_products", 0) or 0 for o in month_orders),
            "total_qty": sum(o.get("total_qty", 0) or 0 for o in month_orders),
        }

        for ci, col_name in enumerate(CALC_COLUMNS, 1):
            val = subtotals.get(col_name)
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _MONTH_SUBTOTAL_FONT
            cell.fill = _MONTH_SUBTOTAL_FILL
            cell.border = _THIN_BORDER
            if col_name in CURRENCY_COLS and isinstance(val, (int, float)):
                cell.number_format = fmt
            elif col_name in PERCENT_COLS and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            if col_name == "profit" and isinstance(val, (int, float)):
                cell.font = Font(name="Calibri", bold=True, size=10,
                                 color="006100" if val >= 0 else "9C0006")

        row += 1  # blank row between months
        row += 1

    logger.info("Sheet 'חישוב': %d orders in %d months", len(orders), len(sorted_months))


def _aggregate_by_field(orders: list[dict[str, Any]], field: str) -> dict[str, dict]:
    """Aggregate orders by a given field."""
    result: dict[str, dict] = {}
    for o in orders:
        key = o.get(field) or "לא ידוע"
        if key not in result:
            result[key] = {"count": 0, "price": 0.0, "vat": 0.0, "cogs": 0.0,
                           "shipping": 0.0, "fees": 0.0, "profit": 0.0}
        result[key]["count"] += 1
        result[key]["price"] += o.get("order_price", 0) or 0
        result[key]["vat"] += o.get("vat", 0) or 0
        result[key]["cogs"] += o.get("cogs", 0) or 0
        result[key]["shipping"] += o.get("shipping", 0) or 0
        result[key]["fees"] += o.get("payment_fee", 0) or 0
        result[key]["profit"] += o.get("profit", 0) or 0
    return result


def _aggregate_monthly(orders: list[dict[str, Any]]) -> dict[str, dict]:
    """Aggregate orders by month (YYYY-MM)."""
    monthly: dict[str, dict] = {}
    for o in orders:
        d = o.get("order_date")
        if d and hasattr(d, "strftime"):
            mk = d.strftime("%Y-%m")
        elif d:
            mk = str(d)[:7]
        else:
            mk = "Unknown"
        if mk not in monthly:
            monthly[mk] = {"count": 0, "price": 0.0, "vat": 0.0, "cogs": 0.0,
                           "shipping": 0.0, "fees": 0.0, "profit": 0.0}
        monthly[mk]["count"] += 1
        monthly[mk]["price"] += o.get("order_price", 0) or 0
        monthly[mk]["vat"] += o.get("vat", 0) or 0
        monthly[mk]["cogs"] += o.get("cogs", 0) or 0
        monthly[mk]["shipping"] += o.get("shipping", 0) or 0
        monthly[mk]["fees"] += o.get("payment_fee", 0) or 0
        monthly[mk]["profit"] += o.get("profit", 0) or 0
    return monthly


def _write_summary_table(
    ws: Worksheet, row: int, title: str, headers: list[str],
    data: dict[str, dict], fmt: str, sort_keys: list[str] | None = None,
    alternating: bool = True,
) -> int:
    """Write a summary section with title, headers, and data rows. Returns next row."""
    _ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    _TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Section title
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    ws.merge_cells(f"A{row}:{get_column_letter(len(headers))}{row}")
    row += 1

    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1

    keys = sort_keys if sort_keys else sorted(data.keys())
    totals = {"count": 0, "price": 0.0, "vat": 0.0, "cogs": 0.0,
              "shipping": 0.0, "fees": 0.0, "profit": 0.0}
    data_start_row = row

    for i, key in enumerate(keys):
        m = data[key]
        margin = m["profit"] / m["price"] if m["price"] else 0
        fill = _ALT_FILL if (alternating and i % 2 == 1) else None

        c = ws.cell(row=row, column=1, value=key)
        c.border = _THIN_BORDER
        c.font = Font(name="Calibri", bold=True, size=10)
        if fill:
            c.fill = fill

        c = ws.cell(row=row, column=2, value=m["count"])
        c.border = _THIN_BORDER
        c.number_format = "#,##0"
        if fill:
            c.fill = fill

        for ci, fld in [(3, "price"), (4, "vat"), (5, "cogs"), (6, "shipping"), (7, "fees"), (8, "profit")]:
            c = ws.cell(row=row, column=ci, value=m[fld])
            c.number_format = fmt
            c.border = _THIN_BORDER
            if fill:
                c.fill = fill
            if fld == "profit":
                c.font = Font(name="Calibri", size=10, bold=True,
                              color="006100" if m[fld] >= 0 else "9C0006")

        c = ws.cell(row=row, column=9, value=margin)
        c.number_format = "0.0%"
        c.border = _THIN_BORDER
        if fill:
            c.fill = fill

        for fld in totals:
            totals[fld] += m[fld]
        row += 1

    # Totals row
    total_margin = totals["profit"] / totals["price"] if totals["price"] else 0
    c = ws.cell(row=row, column=1, value="סה״כ")
    c.font = Font(name="Calibri", bold=True, size=11)
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER

    c = ws.cell(row=row, column=2, value=totals["count"])
    c.font = Font(name="Calibri", bold=True, size=10)
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.number_format = "#,##0"

    for ci, fld in [(3, "price"), (4, "vat"), (5, "cogs"), (6, "shipping"), (7, "fees"), (8, "profit")]:
        c = ws.cell(row=row, column=ci, value=totals[fld])
        c.number_format = fmt
        c.font = Font(name="Calibri", bold=True, size=10,
                      color="006100" if (fld == "profit" and totals[fld] >= 0)
                      else "9C0006" if (fld == "profit") else "000000")
        c.fill = _TOTAL_FILL
        c.border = _THIN_BORDER

    c = ws.cell(row=row, column=9, value=total_margin)
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", bold=True, size=10)
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    row += 1

    return row, data_start_row


def _build_dashboard(wb: Workbook, orders: list[dict[str, Any]], ads_data: dict[str, Any] | None = None) -> None:
    if "דשבורד" in wb.sheetnames:
        del wb["דשבורד"]
    ws = wb.create_sheet("דשבורד")
    ws.sheet_view.rightToLeft = True
    fmt = _currency_fmt()

    if not orders:
        ws.cell(row=1, column=1, value="אין הזמנות להצגה.").font = _KPI_FONT
        return

    priced = [o for o in orders if (o.get("order_price") or 0) > 0]
    total_price = sum(o.get("order_price", 0) or 0 for o in orders)
    total_vat = sum(o.get("vat", 0) or 0 for o in orders)
    total_cogs = sum(o.get("cogs", 0) or 0 for o in orders)
    total_shipping = sum(o.get("shipping", 0) or 0 for o in orders)
    total_fees = sum(o.get("payment_fee", 0) or 0 for o in orders)
    total_profit = sum(o.get("profit", 0) or 0 for o in orders)
    order_count = len(orders)
    priced_count = len(priced)
    avg_margin = (total_profit / total_price) if total_price else 0
    avg_order_value = total_price / priced_count if priced_count else 0
    avg_profit_per_order = total_profit / order_count if order_count else 0
    total_qty = sum(o.get("total_qty", 0) or 0 for o in orders)

    # Ads data
    monthly_ads = ads_data.get("monthly", {}) if ads_data else {}
    ad_spend = monthly_ads.get("spend", 0)
    net_profit = total_profit - ad_spend
    roas = total_price / ad_spend if ad_spend > 0 else 0

    # ── Title + Last Updated ────────────────────────────────────
    row = 1
    ws.merge_cells("A1:I1")
    ws["A1"].value = "OneZoneJersey — דשבורד רווח והפסד"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"עדכון אחרון: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── KPI Cards (2 columns layout) ────────────────────────────
    _KPI_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    _KPI_VALUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    _KPI_PROFIT_POS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _KPI_PROFIT_NEG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    row = 4
    ws.cell(row=row, column=1, value="מדדים מרכזיים").font = _SECTION_FONT
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    kpis_left = [
        ("סה״כ הכנסות", total_price, fmt),
        ("סה״כ עלות סחורה", total_cogs, fmt),
        ("סה״כ מע״מ", total_vat, fmt),
        ("סה״כ משלוחים", total_shipping, fmt),
        ("סה״כ עמלות", total_fees, fmt),
        ("רווח גולמי", total_profit, fmt),
        ("הוצאות פרסום (החודש)", ad_spend, fmt),
        ("רווח נקי (אחרי פרסום)", net_profit, fmt),
    ]
    kpis_right = [
        ("סה״כ הזמנות", order_count, "#,##0"),
        ("סה״כ פריטים שנמכרו", total_qty, "#,##0"),
        ("ממוצע הזמנה", avg_order_value, fmt),
        ("ממוצע רווח להזמנה", avg_profit_per_order, fmt),
        ("אחוז רווח כולל", avg_margin, "0.0%"),
        ("ROAS", roas, "0.00x"),
        ("חשיפות", monthly_ads.get("impressions", 0), "#,##0"),
        ("קליקים", monthly_ads.get("clicks", 0), "#,##0"),
    ]

    for i, ((l_label, l_val, l_fmt), (r_label, r_val, r_fmt)) in enumerate(zip(kpis_left, kpis_right)):
        r = row + i

        # Left KPI (columns A-B)
        lc = ws.cell(row=r, column=1, value=l_label)
        lc.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        lc.fill = _KPI_TITLE_FILL
        lc.border = _THIN_BORDER
        lc.alignment = Alignment(horizontal="right", indent=1)

        vc = ws.cell(row=r, column=2, value=l_val)
        vc.number_format = l_fmt
        vc.border = _THIN_BORDER
        vc.alignment = Alignment(horizontal="center")
        if l_label in ("רווח נקי (אחרי פרסום)", "רווח גולמי"):
            vc.font = Font(name="Calibri", bold=True, size=13,
                           color="006100" if l_val >= 0 else "9C0006")
            vc.fill = _KPI_PROFIT_POS_FILL if l_val >= 0 else _KPI_PROFIT_NEG_FILL
        else:
            vc.font = Font(name="Calibri", bold=True, size=11)
            vc.fill = _KPI_VALUE_FILL

        # Spacer column C
        ws.cell(row=r, column=3, value="")

        # Right KPI (columns D-E)
        lc2 = ws.cell(row=r, column=4, value=r_label)
        lc2.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        lc2.fill = _KPI_TITLE_FILL
        lc2.border = _THIN_BORDER
        lc2.alignment = Alignment(horizontal="right", indent=1)

        vc2 = ws.cell(row=r, column=5, value=r_val)
        vc2.font = Font(name="Calibri", bold=True, size=11)
        vc2.number_format = r_fmt
        vc2.fill = _KPI_VALUE_FILL
        vc2.border = _THIN_BORDER
        vc2.alignment = Alignment(horizontal="center")

    row += len(kpis_left) + 1

    # ── Cost Breakdown Pie Chart ────────────────────────────────
    pie_data_row = row
    ws.cell(row=row, column=7, value="קטגוריה").font = Font(size=1, color="FFFFFF")
    ws.cell(row=row, column=8, value="סכום").font = Font(size=1, color="FFFFFF")
    cost_items = [
        ("רווח", max(total_profit, 0)),
        ("עלות סחורה", total_cogs),
        ("מע״מ", total_vat),
        ("משלוח", total_shipping),
        ("עמלות", total_fees),
    ]
    for i, (cat, val) in enumerate(cost_items):
        ws.cell(row=row + 1 + i, column=7, value=cat).font = Font(size=1, color="FFFFFF")
        ws.cell(row=row + 1 + i, column=8, value=round(val, 0)).font = Font(size=1, color="FFFFFF")

    pie = PieChart()
    pie.title = "פילוח הכנסות"
    pie.style = 10
    labels = Reference(ws, min_col=7, min_row=pie_data_row + 1, max_row=pie_data_row + len(cost_items))
    data_ref = Reference(ws, min_col=8, min_row=pie_data_row, max_row=pie_data_row + len(cost_items))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 16
    pie.height = 12

    # Color the slices
    pie_colors = ["27AE60", "E74C3C", "F39C12", "3498DB", "9B59B6"]
    for idx, color in enumerate(pie_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False

    ws.add_chart(pie, f"F{row - len(kpis_left)}")

    row += len(cost_items) + 2

    # ── Monthly Summary + Charts ────────────────────────────────
    monthly = _aggregate_monthly(orders)
    sorted_months = sorted(monthly.keys())
    # Convert month keys to Hebrew for display
    monthly_heb = {}
    month_key_map = {}  # heb -> original key
    for mk in sorted_months:
        heb = _month_heb(mk)
        monthly_heb[heb] = monthly[mk]
        month_key_map[heb] = mk
    sorted_months_heb = [_month_heb(mk) for mk in sorted_months]

    headers = ["חודש", "הזמנות", "הכנסות", "מע״מ", "עלות סחורה", "משלוח", "עמלות", "רווח", "מרווח"]
    row, monthly_data_start = _write_summary_table(
        ws, row, "סיכום חודשי", headers, monthly_heb, fmt, sort_keys=sorted_months_heb,
    )

    # MoM % change column
    if len(sorted_months) > 1:
        chg_col = 10
        ws.cell(row=monthly_data_start - 1, column=chg_col, value="שינוי %").font = _HEADER_FONT
        ws.cell(row=monthly_data_start - 1, column=chg_col).fill = _HEADER_FILL
        ws.cell(row=monthly_data_start - 1, column=chg_col).border = _THIN_BORDER
        ws.cell(row=monthly_data_start - 1, column=chg_col).alignment = Alignment(horizontal="center")

        prev_profit = None
        for i, mk in enumerate(sorted_months):
            cur_profit = monthly[mk]["profit"]
            r = monthly_data_start + i
            if prev_profit is not None and prev_profit != 0:
                change = (cur_profit - prev_profit) / abs(prev_profit)
                c = ws.cell(row=r, column=chg_col, value=change)
                c.number_format = "+0.0%;-0.0%"
                c.border = _THIN_BORDER
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="006100" if change >= 0 else "9C0006")
            else:
                ws.cell(row=r, column=chg_col, value="—").border = _THIN_BORDER
            prev_profit = cur_profit

    # Line chart — Monthly Profit Trend
    if len(sorted_months) >= 2:
        line = LineChart()
        line.title = "טרנד רווח חודשי"
        line.y_axis.title = "רווח (₪)" if settings.CURRENCY == "ILS" else "רווח ($)"
        line.x_axis.title = "חודש"
        line.style = 10
        line.width = 24
        line.height = 13

        profit_ref = Reference(ws, min_col=8, min_row=monthly_data_start - 1,
                               max_row=monthly_data_start + len(sorted_months) - 1)
        cats = Reference(ws, min_col=1, min_row=monthly_data_start,
                         max_row=monthly_data_start + len(sorted_months) - 1)
        line.add_data(profit_ref, titles_from_data=True)
        line.set_categories(cats)
        line.series[0].graphicalProperties.line.solidFill = "27AE60"
        line.series[0].graphicalProperties.line.width = 28000
        line.series[0].smooth = True

        ws.add_chart(line, f"A{row + 1}")

    # Bar chart — Monthly Revenue vs Costs
    if len(sorted_months) >= 2:
        bar = BarChart()
        bar.type = "col"
        bar.title = "הכנסות מול עלויות לפי חודש"
        bar.style = 10
        bar.width = 24
        bar.height = 13

        rev_ref = Reference(ws, min_col=3, min_row=monthly_data_start - 1,
                            max_row=monthly_data_start + len(sorted_months) - 1)
        cogs_ref = Reference(ws, min_col=5, min_row=monthly_data_start - 1,
                             max_row=monthly_data_start + len(sorted_months) - 1)
        profit_ref2 = Reference(ws, min_col=8, min_row=monthly_data_start - 1,
                                max_row=monthly_data_start + len(sorted_months) - 1)
        cats = Reference(ws, min_col=1, min_row=monthly_data_start,
                         max_row=monthly_data_start + len(sorted_months) - 1)

        bar.add_data(rev_ref, titles_from_data=True)
        bar.add_data(cogs_ref, titles_from_data=True)
        bar.add_data(profit_ref2, titles_from_data=True)
        bar.set_categories(cats)

        bar_colors = ["3498DB", "E74C3C", "27AE60"]
        for i, color in enumerate(bar_colors):
            if i < len(bar.series):
                bar.series[i].graphicalProperties.solidFill = color

        ws.add_chart(bar, f"F{row + 1}")

    row += 17  # space for charts

    # ── Status Summary ──────────────────────────────────────────
    statuses = _aggregate_by_field(orders, "status")
    headers2 = ["סטטוס", "הזמנות", "הכנסות", "מע״מ", "עלות סחורה", "משלוח", "עמלות", "רווח", "מרווח"]
    row, _ = _write_summary_table(ws, row, "סיכום לפי סטטוס", headers2, statuses, fmt)

    row += 1

    # ── Shipping Type Summary ───────────────────────────────────
    ship_types = _aggregate_by_field(orders, "shipping_type")
    headers3 = ["סוג משלוח", "הזמנות", "הכנסות", "מע״מ", "עלות סחורה", "עלות משלוח", "עמלות", "רווח", "מרווח"]
    row, _ = _write_summary_table(ws, row, "סיכום לפי סוג משלוח", headers3, ship_types, fmt)

    row += 1

    # ── Product Type Summary ────────────────────────────────────
    product_types = _aggregate_by_field(orders, "product_type")
    headers4 = ["סוג מוצר", "הזמנות", "הכנסות", "מע״מ", "עלות סחורה", "משלוח", "עמלות", "רווח", "מרווח"]
    row, _ = _write_summary_table(ws, row, "סיכום לפי סוג מוצר", headers4, product_types, fmt)

    row += 1

    # ── Top 15 Customers ────────────────────────────────────────
    customers = _aggregate_by_field(orders, "customer_name")
    top_customers = dict(sorted(customers.items(), key=lambda x: x[1]["price"], reverse=True)[:15])
    headers5 = ["לקוח", "הזמנות", "הכנסות", "מע״מ", "עלות סחורה", "משלוח", "עמלות", "רווח", "מרווח"]
    top_keys = list(sorted(top_customers.keys(), key=lambda k: top_customers[k]["price"], reverse=True))
    row, _ = _write_summary_table(ws, row, "15 לקוחות מובילים (לפי הכנסות)", headers5, top_customers, fmt,
                                  sort_keys=top_keys)

    # ── Column widths ───────────────────────────────────────────
    widths = {"A": 28, "B": 14, "C": 16, "D": 16, "E": 16, "F": 14, "G": 14, "H": 16, "I": 12, "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Hide gridlines
    ws.sheet_view.showGridLines = False


# ── Public API ──────────────────────────────────────────────────

def _build_weekly_summary(wb: Workbook, orders: list[dict[str, Any]], ads_data: dict[str, Any]) -> None:
    """Build weekly summary sheet for the current month with orders + ads + ROAS."""
    if "סיכום שבועי" in wb.sheetnames:
        del wb["סיכום שבועי"]
    ws = wb.create_sheet("סיכום שבועי")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    fmt = _currency_fmt()

    from datetime import date
    today = date.today()
    month_key = today.strftime("%Y-%m")
    month_name = _month_heb(month_key)

    # Filter orders for current month
    month_orders = [o for o in orders
                    if o.get("order_date") and hasattr(o["order_date"], "strftime")
                    and o["order_date"].strftime("%Y-%m") == month_key]

    weekly_ads = ads_data.get("weekly", []) if ads_data else []
    monthly_ads = ads_data.get("monthly", {}) if ads_data else {}

    # Title
    row = 1
    ws.merge_cells("A1:J1")
    ws["A1"].value = f"OneZoneJersey — סיכום שבועי — {month_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"עדכון אחרון: {today.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Monthly KPIs
    row = 4
    ws.cell(row=row, column=1, value="סיכום חודשי").font = _SECTION_FONT
    ws.merge_cells(f"A{row}:J{row}")
    row += 1

    total_revenue = sum(o.get("order_price", 0) or 0 for o in month_orders)
    total_profit = sum(o.get("profit", 0) or 0 for o in month_orders)
    total_spend = monthly_ads.get("spend", 0)
    net_profit = total_profit - total_spend
    roas = total_revenue / total_spend if total_spend > 0 else 0
    net_margin = net_profit / total_revenue * 100 if total_revenue > 0 else 0

    _KPI_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    _KPI_VALUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    _KPI_PROFIT_POS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _KPI_PROFIT_NEG = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    kpis = [
        ("סה״כ הכנסות", total_revenue, fmt, None),
        ("סה״כ הוצאות פרסום", total_spend, fmt, None),
        ("רווח גולמי (לפני פרסום)", total_profit, fmt, None),
        ("רווח נקי (אחרי פרסום)", net_profit, fmt, "profit"),
        ("ROAS", roas, "0.00x", None),
        ("מרווח נקי", net_margin / 100, "0.0%", None),
        ("הזמנות", len(month_orders), "#,##0", None),
        ("חשיפות", monthly_ads.get("impressions", 0), "#,##0", None),
        ("קליקים", monthly_ads.get("clicks", 0), "#,##0", None),
    ]

    col_left = 1
    for i, (label, value, nfmt, special) in enumerate(kpis):
        r = row + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        lc.fill = _KPI_TITLE_FILL
        lc.border = _THIN_BORDER
        lc.alignment = Alignment(horizontal="right", indent=1)

        vc = ws.cell(row=r, column=2, value=value)
        vc.number_format = nfmt
        vc.border = _THIN_BORDER
        vc.alignment = Alignment(horizontal="center")
        if special == "profit":
            vc.font = Font(name="Calibri", bold=True, size=13,
                           color="006100" if value >= 0 else "9C0006")
            vc.fill = _KPI_PROFIT_POS if value >= 0 else _KPI_PROFIT_NEG
        else:
            vc.font = Font(name="Calibri", bold=True, size=11)
            vc.fill = _KPI_VALUE_FILL

    row += len(kpis) + 2

    # Weekly breakdown table
    ws.cell(row=row, column=1, value="פירוט שבועי").font = _SECTION_FONT
    ws.merge_cells(f"A{row}:J{row}")
    row += 1

    headers = ["שבוע", "תאריכים", "ימים", "הזמנות", "הכנסות", "רווח גולמי",
               "הוצאות פרסום", "רווח נקי", "ROAS", "מרווח נקי"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1

    _ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    _TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    grand = {"orders": 0, "revenue": 0, "profit": 0, "spend": 0}

    for i, week in enumerate(weekly_ads):
        fill = _ALT_FILL if i % 2 == 1 else None
        w_start = date.fromisoformat(week["start"])
        w_end = date.fromisoformat(week["end"])

        # Orders for this week
        w_orders = [o for o in month_orders
                    if o.get("order_date") and w_start <= o["order_date"] <= w_end]
        w_revenue = sum(o.get("order_price", 0) or 0 for o in w_orders)
        w_profit = sum(o.get("profit", 0) or 0 for o in w_orders)
        w_spend = week["spend"]
        w_net = w_profit - w_spend
        w_roas = w_revenue / w_spend if w_spend > 0 else 0
        w_margin = w_net / w_revenue if w_revenue > 0 else 0

        grand["orders"] += len(w_orders)
        grand["revenue"] += w_revenue
        grand["profit"] += w_profit
        grand["spend"] += w_spend

        values = [
            (f"שבוע {week['week']}", None, None),
            (f"{week['start_display']} - {week['end_display']}", None, None),
            (week["days"], "#,##0", None),
            (len(w_orders), "#,##0", None),
            (w_revenue, fmt, None),
            (w_profit, fmt, None),
            (w_spend, fmt, None),
            (w_net, fmt, "profit"),
            (w_roas, "0.00x", None),
            (w_margin, "0.0%", None),
        ]

        for ci, (val, nfmt, special) in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = _THIN_BORDER
            if fill:
                c.fill = fill
            if nfmt:
                c.number_format = nfmt
            if special == "profit":
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="006100" if val >= 0 else "9C0006")
            else:
                c.font = Font(name="Calibri", size=10)

        row += 1

    # Totals row
    g_net = grand["profit"] - grand["spend"]
    g_roas = grand["revenue"] / grand["spend"] if grand["spend"] > 0 else 0
    g_margin = g_net / grand["revenue"] if grand["revenue"] > 0 else 0

    totals = [
        ("סה״כ", None), ("", None), ("", None),
        (grand["orders"], "#,##0"), (grand["revenue"], fmt), (grand["profit"], fmt),
        (grand["spend"], fmt), (g_net, fmt), (g_roas, "0.00x"), (g_margin, "0.0%"),
    ]
    for ci, (val, nfmt) in enumerate(totals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = Font(name="Calibri", bold=True, size=11)
        c.fill = _TOTAL_FILL
        c.border = _THIN_BORDER
        if nfmt:
            c.number_format = nfmt
        if ci == 8:  # net profit
            c.font = Font(name="Calibri", bold=True, size=11,
                          color="006100" if val >= 0 else "9C0006")

    # Column widths
    widths = {"A": 18, "B": 18, "C": 8, "D": 12, "E": 16, "F": 16,
              "G": 16, "H": 16, "I": 10, "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_workbook(
    raw_orders: list[dict[str, Any]],
    calc_orders: list[dict[str, Any]],
    output_path: str | None = None,
    ads_data: dict[str, Any] | None = None,
) -> Path:
    path = Path(output_path or settings.OUTPUT_FILE)

    if path.exists():
        logger.info("Loading existing workbook: %s", path)
        wb = load_workbook(path)
    else:
        logger.info("Creating new workbook: %s", path)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Remove old English sheet names if they exist
    for old_name in ["Settings", "Orders_Raw", "Calc", "Dashboard"]:
        if old_name in wb.sheetnames:
            del wb[old_name]

    _build_settings_sheet(wb)
    _build_raw_sheet(wb, raw_orders)
    _build_calc_sheet(wb, calc_orders)
    _build_dashboard(wb, calc_orders, ads_data=ads_data)
    _build_weekly_summary(wb, calc_orders, ads_data or {})

    # Reorder sheets
    desired = ["הגדרות", "הזמנות_גולמי", "חישוב", "דשבורד", "סיכום שבועי"]
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    # Validate no duplicates
    _validate_no_duplicates(wb["חישוב"])

    wb.save(path)
    logger.info("Workbook saved: %s (%d orders)", path, len(calc_orders))
    return path


def _validate_no_duplicates(ws: Worksheet) -> None:
    """Check for duplicate order_ids (skip month headers and subtotal rows)."""
    seen: set[str] = set()
    dupes: list[str] = []
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=1).value
        if val is None:
            continue
        key = str(val)
        # Skip month headers and subtotal rows
        if "|" in key or key.startswith("סה״כ") or key.startswith("TOTAL") or len(key) < 4:
            continue
        if key in seen:
            dupes.append(key)
        seen.add(key)
    if dupes:
        logger.warning("Duplicate order_ids in חישוב: %s", dupes)
    else:
        logger.info("No duplicate order_ids in Calc sheet.")
