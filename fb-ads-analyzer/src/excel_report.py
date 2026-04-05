"""
Excel report generator — comprehensive Hebrew RTL ads analysis workbook.
Sheets: דשבורד, פאנל, קמפיינים, מודעות, יומי, שבועי, חודשי, קהלים, פלטפורמות, המלצות
"""
import logging
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import settings

logger = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
NEUTRAL_FILL = PatternFill("solid", fgColor="D9E2F3")
KPI_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=13, color="2F5496")
MONEY_FMT2 = '₪#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Priority colors for recommendations
PRIORITY_FILLS = {
    1: PatternFill("solid", fgColor="FFC7CE"),  # Red - urgent
    2: PatternFill("solid", fgColor="FFEB9C"),  # Yellow - important
    3: PatternFill("solid", fgColor="C6EFCE"),  # Green - opportunity
    4: PatternFill("solid", fgColor="D9E2F3"),  # Blue - info
    5: PatternFill("solid", fgColor="E2EFDA"),  # Light green - ok
}


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data(ws, start_row, end_row, cols, money_cols=None, pct_cols=None):
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in money_cols:
                cell.number_format = MONEY_FMT2
            elif c in pct_cols:
                cell.number_format = '0.00'


def _auto_width(ws, cols, min_width=12):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min_width


def _inf_safe(val):
    """Convert inf to 0 for Excel."""
    return 0 if val == float("inf") else val


# ═══════════════════════════════════════════════════════════════
# SHEET 0: דשבורד (Dashboard)
# ═══════════════════════════════════════════════════════════════

def _build_dashboard(wb: Workbook, campaign_data: dict, trend_data: dict):
    ws = wb.create_sheet("דשבורד", 0)
    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value="OneZoneJersey — ניתוח מודעות מעמיק").font = Font(
        name="Arial", bold=True, size=20, color="1F3864")
    ws.cell(row=2, column=1, value=f"עדכון: {date.today().strftime('%d/%m/%Y')}").font = Font(
        size=11, color="888888")

    total_spend = campaign_data.get("total_spend", 0)
    total_purchases = campaign_data.get("total_purchases", 0)
    total_pv = campaign_data.get("total_purchase_value", 0)
    total_roas = campaign_data.get("total_roas", 0)
    total_atc = campaign_data.get("total_atc", 0)
    total_ic = campaign_data.get("total_ic", 0)
    cpp = total_spend / total_purchases if total_purchases > 0 else 0
    avg_freq = campaign_data.get("avg_frequency", 0)

    trend_map = {"improving": "📈 עלייה", "declining": "📉 ירידה", "stable": "➡️ יציב"}

    kpis = [
        ("הוצאה כוללת", f"₪{total_spend:,.0f}"),
        ("הכנסות", f"₪{total_pv:,.0f}"),
        ("ROAS", f"{total_roas:.2f}x"),
        ("רכישות", f"{total_purchases:,}"),
        ("עלות/רכישה", f"₪{cpp:,.0f}"),
        ("הוספות לעגלה", f"{total_atc:,}"),
        ("התחלת תשלום", f"{total_ic:,}"),
        ("תדירות ממוצעת", f"{avg_freq:.1f}"),
        ("מגמה", trend_map.get(trend_data.get("trend", ""), "—")),
    ]

    row = 4
    for i, (label, value) in enumerate(kpis):
        col = 1 + (i % 3) * 3
        if i > 0 and i % 3 == 0:
            row += 3
        ws.cell(row=row, column=col, value=label).font = Font(name="Arial", size=10, color="666666")
        ws.cell(row=row + 1, column=col, value=value).font = KPI_FONT
        for r in range(row, row + 2):
            for c in range(col, col + 2):
                ws.cell(row=r, column=c).fill = NEUTRAL_FILL
                ws.cell(row=r, column=c).border = THIN_BORDER

    # ── Funnel summary ──
    funnel_row = row + 4
    ws.cell(row=funnel_row, column=1, value="פאנל המרה").font = SUBTITLE_FONT
    total_clicks = campaign_data.get("total_clicks", 0)
    funnel_items = [
        ("קליקים", total_clicks, "—"),
        ("הוספה לעגלה", total_atc,
         f"{total_atc / total_clicks * 100:.1f}%" if total_clicks > 0 else "—"),
        ("התחלת תשלום", total_ic,
         f"{total_ic / total_atc * 100:.1f}%" if total_atc > 0 else "—"),
        ("רכישות", total_purchases,
         f"{total_purchases / total_ic * 100:.1f}%" if total_ic > 0 else "—"),
    ]
    for i, (label, val, rate) in enumerate(funnel_items):
        r = funnel_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=11)
        ws.cell(row=r, column=2, value=val).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=r, column=3, value=rate).font = Font(name="Arial", size=11, color="666666")
        if i > 0:
            ws.cell(row=r, column=4, value="→").font = Font(size=14)

    _auto_width(ws, 9, 14)


# ═══════════════════════════════════════════════════════════════
# SHEET 1: קמפיינים (Campaigns)
# ═══════════════════════════════════════════════════════════════

def _build_campaigns_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("קמפיינים")
    ws.sheet_view.rightToLeft = True

    headers = ["קמפיין", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים", "CTR%",
               "הוספה לעגלה", "התחלת תשלום", "רכישות", "עלות/רכישה",
               "% קליק→עגלה", "% עגלה→תשלום", "% תשלום→רכישה", "תדירות"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    campaigns = data.get("campaigns", [])
    for i, camp in enumerate(campaigns, 2):
        ws.cell(row=i, column=1, value=camp.get("campaign_name", ""))
        ws.cell(row=i, column=2, value=camp["spend"])
        ws.cell(row=i, column=3, value=camp["purchase_value"])
        ws.cell(row=i, column=4, value=camp["roas"])
        ws.cell(row=i, column=5, value=camp["impressions"])
        ws.cell(row=i, column=6, value=camp["clicks"])
        ws.cell(row=i, column=7, value=camp["ctr"])
        ws.cell(row=i, column=8, value=camp["add_to_cart"])
        ws.cell(row=i, column=9, value=camp["initiate_checkout"])
        ws.cell(row=i, column=10, value=camp["purchases"])
        ws.cell(row=i, column=11, value=_inf_safe(camp["cost_per_result"]))
        ws.cell(row=i, column=12, value=camp.get("atc_rate", 0))
        ws.cell(row=i, column=13, value=camp.get("ic_rate", 0))
        ws.cell(row=i, column=14, value=camp.get("purchase_rate", 0))
        ws.cell(row=i, column=15, value=camp.get("frequency", 0))

        # Color: ROAS
        roas_cell = ws.cell(row=i, column=4)
        if camp["roas"] >= 2:
            roas_cell.fill = GOOD_FILL
        elif camp["roas"] > 0:
            roas_cell.fill = WARN_FILL
        elif camp["spend"] > 0:
            roas_cell.fill = BAD_FILL

        # Color: frequency
        if camp.get("frequency", 0) > 4:
            ws.cell(row=i, column=15).fill = BAD_FILL

    end = 1 + len(campaigns)
    _style_data(ws, 2, end, len(headers),
                money_cols=[2, 3, 11], pct_cols=[7, 12, 13, 14])
    _auto_width(ws, len(headers), 13)
    ws.column_dimensions["A"].width = 35


# ── Group styles for ads sheet ──
CAMPAIGN_FILL = PatternFill("solid", fgColor="1F3864")
CAMPAIGN_FONT = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ADSET_FILL = PatternFill("solid", fgColor="4472C4")
ADSET_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")


# ═══════════════════════════════════════════════════════════════
# SHEET 2: מודעות (Ads — grouped by Campaign → Ad Set → Ad)
# ═══════════════════════════════════════════════════════════════

def _build_ads_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("מודעות")
    ws.sheet_view.rightToLeft = True

    # Data headers for ad rows
    ad_headers = ["מודעה", "הוצאה", "הכנסות", "ROAS",
                  "חשיפות", "קליקים", "CTR%",
                  "הוספה לעגלה", "התחלת תשלום", "רכישות",
                  "עלות/רכישה", "עלות/עגלה", "% המרה כוללת", "תדירות"]
    num_cols = len(ad_headers)

    # ── Group ads by campaign → adset ──
    ads = data.get("ads", [])
    from collections import OrderedDict
    tree: OrderedDict[str, OrderedDict[str, list]] = OrderedDict()
    for ad in ads:
        camp = ad.get("campaign_name", "ללא קמפיין")
        aset = ad.get("adset_name", "ללא קבוצה")
        tree.setdefault(camp, OrderedDict()).setdefault(aset, []).append(ad)

    row = 1

    for camp_name, adsets in tree.items():
        # ── Campaign header row ──
        camp_ads = [ad for aset_ads in adsets.values() for ad in aset_ads]
        camp_spend = sum(a["spend"] for a in camp_ads)
        camp_pv = sum(a["purchase_value"] for a in camp_ads)
        camp_purch = sum(a["purchases"] for a in camp_ads)
        camp_atc = sum(a["add_to_cart"] for a in camp_ads)
        camp_roas = camp_pv / camp_spend if camp_spend > 0 else 0

        ws.cell(row=row, column=1,
                value=f"📁 קמפיין: {camp_name}")
        ws.cell(row=row, column=4,
                value=f"הוצאה: ₪{camp_spend:,.0f}")
        ws.cell(row=row, column=7,
                value=f"ROAS: {camp_roas:.2f}")
        ws.cell(row=row, column=10,
                value=f"רכישות: {camp_purch}")
        ws.cell(row=row, column=12,
                value=f"עגלה: {camp_atc}")
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = CAMPAIGN_FILL
            cell.font = CAMPAIGN_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for aset_name, aset_ads in adsets.items():
            # ── Ad Set sub-header row ──
            aset_spend = sum(a["spend"] for a in aset_ads)
            aset_pv = sum(a["purchase_value"] for a in aset_ads)
            aset_purch = sum(a["purchases"] for a in aset_ads)
            aset_atc = sum(a["add_to_cart"] for a in aset_ads)
            aset_ic = sum(a["initiate_checkout"] for a in aset_ads)
            aset_roas = aset_pv / aset_spend if aset_spend > 0 else 0

            ws.cell(row=row, column=1,
                    value=f"   📂 Ad Set: {aset_name}")
            ws.cell(row=row, column=4,
                    value=f"₪{aset_spend:,.0f}")
            ws.cell(row=row, column=7,
                    value=f"ROAS: {aset_roas:.2f}")
            ws.cell(row=row, column=10,
                    value=f"{aset_purch} רכישות")
            ws.cell(row=row, column=8,
                    value=f"{aset_atc} עגלה")
            ws.cell(row=row, column=9,
                    value=f"{aset_ic} תשלום")
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = ADSET_FILL
                cell.font = ADSET_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 25
            row += 1

            # ── Column headers for ads under this ad set ──
            for c, h in enumerate(ad_headers, 1):
                ws.cell(row=row, column=c, value=h)
            _style_header(ws, row, num_cols)
            row += 1

            # ── Ad rows ──
            for ad in aset_ads:
                ws.cell(row=row, column=1, value=ad.get("ad_name", ""))
                ws.cell(row=row, column=2, value=ad["spend"])
                ws.cell(row=row, column=3, value=ad["purchase_value"])
                ws.cell(row=row, column=4, value=ad.get("roas", 0))
                ws.cell(row=row, column=5, value=ad["impressions"])
                ws.cell(row=row, column=6, value=ad["clicks"])
                ws.cell(row=row, column=7, value=ad["ctr"])
                ws.cell(row=row, column=8, value=ad["add_to_cart"])
                ws.cell(row=row, column=9, value=ad["initiate_checkout"])
                ws.cell(row=row, column=10, value=ad["purchases"])
                ws.cell(row=row, column=11, value=_inf_safe(ad["cost_per_result"]))
                ws.cell(row=row, column=12, value=_inf_safe(ad.get("cost_per_atc", 0)))
                ws.cell(row=row, column=13, value=ad.get("conv_rate", 0))
                ws.cell(row=row, column=14, value=ad.get("frequency", 0))

                # Style data cells
                for c in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if c in [2, 3, 11, 12]:
                        cell.number_format = MONEY_FMT2
                    elif c in [7, 13]:
                        cell.number_format = '0.00'

                # Color ROAS
                roas_val = ad.get("roas", 0)
                if roas_val >= 2:
                    ws.cell(row=row, column=4).fill = GOOD_FILL
                elif roas_val > 0:
                    ws.cell(row=row, column=4).fill = WARN_FILL
                elif ad["spend"] > 50 and ad["purchases"] == 0:
                    ws.cell(row=row, column=4).fill = BAD_FILL

                # Color frequency
                if ad.get("frequency", 0) > 5:
                    ws.cell(row=row, column=14).fill = BAD_FILL

                row += 1

            # Spacer row between ad sets
            row += 1

    _auto_width(ws, num_cols, 12)
    ws.column_dimensions["A"].width = 35


# ═══════════════════════════════════════════════════════════════
# SHEET 3: יומי (Daily Trends)
# ═══════════════════════════════════════════════════════════════

def _build_daily_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("יומי")
    ws.sheet_view.rightToLeft = True

    headers = ["תאריך", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים", "CTR%",
               "הוספה לעגלה", "התחלת תשלום", "רכישות", "עלות/רכישה", "תדירות"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    daily = data.get("daily", [])
    for i, d in enumerate(daily, 2):
        ws.cell(row=i, column=1, value=d["date"])
        ws.cell(row=i, column=2, value=d["spend"])
        ws.cell(row=i, column=3, value=d.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=d.get("roas", 0))
        ws.cell(row=i, column=5, value=d["impressions"])
        ws.cell(row=i, column=6, value=d["clicks"])
        ws.cell(row=i, column=7, value=d["ctr"])
        ws.cell(row=i, column=8, value=d.get("add_to_cart", 0))
        ws.cell(row=i, column=9, value=d.get("initiate_checkout", 0))
        ws.cell(row=i, column=10, value=d["purchases"])
        ws.cell(row=i, column=11, value=d["cost_per_purchase"])
        ws.cell(row=i, column=12, value=d.get("frequency", 0))

    end = 1 + len(daily)
    _style_data(ws, 2, end, len(headers), money_cols=[2, 3, 11], pct_cols=[7])
    _auto_width(ws, len(headers))

    if len(daily) > 1:
        cats = Reference(ws, min_col=1, min_row=2, max_row=end)

        # ROAS trend
        chart1 = LineChart()
        chart1.title = "ROAS יומי"
        chart1.y_axis.title = "ROAS"
        chart1.width = 28
        chart1.height = 14
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=end, max_col=4)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.series[0].graphicalProperties.line.width = 25000
        ws.add_chart(chart1, "N2")

        # Purchases bar
        chart2 = BarChart()
        chart2.title = "רכישות יומיות"
        chart2.width = 28
        chart2.height = 14
        data_ref2 = Reference(ws, min_col=10, min_row=1, max_row=end, max_col=10)
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, "N18")

        # Funnel: ATC / IC / Purchases
        chart3 = LineChart()
        chart3.title = "פאנל יומי — עגלה / תשלום / רכישות"
        chart3.width = 28
        chart3.height = 14
        for col in [8, 9, 10]:
            ref = Reference(ws, min_col=col, min_row=1, max_row=end, max_col=col)
            chart3.add_data(ref, titles_from_data=True)
        chart3.set_categories(cats)
        ws.add_chart(chart3, "N34")


# ═══════════════════════════════════════════════════════════════
# SHEET 4: שבועי (Weekly Trends)
# ═══════════════════════════════════════════════════════════════

def _build_weekly_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("שבועי")
    ws.sheet_view.rightToLeft = True

    headers = ["שבוע (תחילה)", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים",
               "הוספה לעגלה", "התחלת תשלום", "רכישות", "עלות/רכישה",
               "% עגלה→רכישה", "תדירות"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    weekly = data.get("weekly", [])
    for i, w in enumerate(weekly, 2):
        ws.cell(row=i, column=1, value=w["week_start"])
        ws.cell(row=i, column=2, value=w["spend"])
        ws.cell(row=i, column=3, value=w.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=w.get("roas", 0))
        ws.cell(row=i, column=5, value=w["impressions"])
        ws.cell(row=i, column=6, value=w["clicks"])
        ws.cell(row=i, column=7, value=w.get("add_to_cart", 0))
        ws.cell(row=i, column=8, value=w.get("initiate_checkout", 0))
        ws.cell(row=i, column=9, value=w["purchases"])
        ws.cell(row=i, column=10, value=w["cost_per_purchase"])
        ws.cell(row=i, column=11, value=w.get("atc_to_purchase", 0))
        ws.cell(row=i, column=12, value=w.get("frequency", 0))

        # Color ROAS
        if w.get("roas", 0) >= 2:
            ws.cell(row=i, column=4).fill = GOOD_FILL
        elif w.get("roas", 0) > 0:
            ws.cell(row=i, column=4).fill = WARN_FILL

    end = 1 + len(weekly)
    _style_data(ws, 2, end, len(headers), money_cols=[2, 3, 10], pct_cols=[11])
    _auto_width(ws, len(headers), 14)

    if len(weekly) > 1:
        cats = Reference(ws, min_col=1, min_row=2, max_row=end)
        chart = BarChart()
        chart.title = "ROAS שבועי"
        chart.width = 24
        chart.height = 14
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=end, max_col=4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "N2")


# ═══════════════════════════════════════════════════════════════
# SHEET 5: חודשי (Monthly Trends)
# ═══════════════════════════════════════════════════════════════

def _build_monthly_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("חודשי")
    ws.sheet_view.rightToLeft = True

    headers = ["חודש", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים",
               "הוספה לעגלה", "התחלת תשלום", "רכישות", "עלות/רכישה",
               "% עגלה→רכישה", "תדירות"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    monthly = data.get("monthly", [])
    for i, m in enumerate(monthly, 2):
        ws.cell(row=i, column=1, value=m["month"])
        ws.cell(row=i, column=2, value=m["spend"])
        ws.cell(row=i, column=3, value=m.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=m.get("roas", 0))
        ws.cell(row=i, column=5, value=m["impressions"])
        ws.cell(row=i, column=6, value=m["clicks"])
        ws.cell(row=i, column=7, value=m.get("add_to_cart", 0))
        ws.cell(row=i, column=8, value=m.get("initiate_checkout", 0))
        ws.cell(row=i, column=9, value=m["purchases"])
        ws.cell(row=i, column=10, value=m["cost_per_purchase"])
        ws.cell(row=i, column=11, value=m.get("atc_to_purchase", 0))
        ws.cell(row=i, column=12, value=m.get("frequency", 0))

        if m.get("roas", 0) >= 2:
            ws.cell(row=i, column=4).fill = GOOD_FILL
        elif m.get("roas", 0) > 0:
            ws.cell(row=i, column=4).fill = WARN_FILL

    end = 1 + len(monthly)
    _style_data(ws, 2, end, len(headers), money_cols=[2, 3, 10], pct_cols=[11])
    _auto_width(ws, len(headers), 14)

    if len(monthly) > 1:
        cats = Reference(ws, min_col=1, min_row=2, max_row=end)
        chart = BarChart()
        chart.title = "ROAS חודשי"
        chart.width = 24
        chart.height = 14
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=end, max_col=4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "N2")

        # Spend vs Revenue
        chart2 = BarChart()
        chart2.title = "הוצאה מול הכנסות — חודשי"
        chart2.width = 24
        chart2.height = 14
        for col in [2, 3]:
            ref = Reference(ws, min_col=col, min_row=1, max_row=end, max_col=col)
            chart2.add_data(ref, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, "N18")


# ═══════════════════════════════════════════════════════════════
# SHEET 6: קהלים (Demographics)
# ═══════════════════════════════════════════════════════════════

def _build_demographics_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("קהלים")
    ws.sheet_view.rightToLeft = True

    # ── By Age ──
    ws.cell(row=1, column=1, value="ניתוח לפי גיל").font = TITLE_FONT
    age_headers = ["גיל", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים", "CTR%",
                   "הוספה לעגלה", "רכישות", "עלות/רכישה"]
    for c, h in enumerate(age_headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header(ws, 2, len(age_headers))

    ages = data.get("by_age", [])
    for i, a in enumerate(ages, 3):
        ws.cell(row=i, column=1, value=a["age"])
        ws.cell(row=i, column=2, value=a["spend"])
        ws.cell(row=i, column=3, value=a.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=a.get("roas", 0))
        ws.cell(row=i, column=5, value=a["impressions"])
        ws.cell(row=i, column=6, value=a["clicks"])
        ws.cell(row=i, column=7, value=a["ctr"])
        ws.cell(row=i, column=8, value=a.get("add_to_cart", 0))
        ws.cell(row=i, column=9, value=a["purchases"])
        ws.cell(row=i, column=10, value=_inf_safe(a["cost_per_purchase"]))

    age_end = 2 + len(ages)
    _style_data(ws, 3, age_end, len(age_headers), money_cols=[2, 3, 10], pct_cols=[7])

    if ages:
        pie = PieChart()
        pie.title = "חלוקת הוצאה לפי גיל"
        pie.width = 18
        pie.height = 14
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=age_end)
        cats = Reference(ws, min_col=1, min_row=3, max_row=age_end)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats)
        ws.add_chart(pie, "L2")

    # ── By Gender ──
    gap = age_end + 3
    ws.cell(row=gap, column=1, value="ניתוח לפי מגדר").font = TITLE_FONT
    gender_headers = ["מגדר", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים",
                      "הוספה לעגלה", "רכישות", "עלות/רכישה"]
    for c, h in enumerate(gender_headers, 1):
        ws.cell(row=gap + 1, column=c, value=h)
    _style_header(ws, gap + 1, len(gender_headers))

    genders = data.get("by_gender", [])
    gender_map = {"male": "גברים", "female": "נשים", "unknown": "לא ידוע"}
    for i, g in enumerate(genders, gap + 2):
        ws.cell(row=i, column=1, value=gender_map.get(g["gender"], g["gender"]))
        ws.cell(row=i, column=2, value=g["spend"])
        ws.cell(row=i, column=3, value=g.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=g.get("roas", 0))
        ws.cell(row=i, column=5, value=g["impressions"])
        ws.cell(row=i, column=6, value=g["clicks"])
        ws.cell(row=i, column=7, value=g.get("add_to_cart", 0))
        ws.cell(row=i, column=8, value=g["purchases"])
        ws.cell(row=i, column=9, value=_inf_safe(g["cost_per_purchase"]))

    gender_end = gap + 1 + len(genders)
    _style_data(ws, gap + 2, gender_end, len(gender_headers), money_cols=[2, 3, 9])
    _auto_width(ws, max(len(age_headers), len(gender_headers)), 14)


# ═══════════════════════════════════════════════════════════════
# SHEET 7: פלטפורמות (Placements)
# ═══════════════════════════════════════════════════════════════

def _build_placements_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("פלטפורמות")
    ws.sheet_view.rightToLeft = True

    headers = ["פלטפורמה", "הוצאה", "הכנסות", "ROAS", "חשיפות", "קליקים", "CTR%",
               "רכישות", "עלות/רכישה"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    placements = data.get("placements", [])
    platform_map = {"facebook": "Facebook", "instagram": "Instagram",
                    "audience_network": "Audience Network", "messenger": "Messenger"}
    for i, p in enumerate(placements, 2):
        ws.cell(row=i, column=1, value=platform_map.get(p.get("platform", ""), p.get("platform", "")))
        ws.cell(row=i, column=2, value=p["spend"])
        ws.cell(row=i, column=3, value=p.get("purchase_value", 0))
        ws.cell(row=i, column=4, value=p.get("roas", 0))
        ws.cell(row=i, column=5, value=p["impressions"])
        ws.cell(row=i, column=6, value=p["clicks"])
        ws.cell(row=i, column=7, value=p.get("ctr", 0))
        ws.cell(row=i, column=8, value=p["purchases"])
        ws.cell(row=i, column=9, value=_inf_safe(p["cost_per_result"]))

    end = 1 + len(placements)
    _style_data(ws, 2, end, len(headers), money_cols=[2, 3, 9], pct_cols=[7])
    _auto_width(ws, len(headers), 16)

    if placements:
        bar = BarChart()
        bar.title = "ROAS לפי פלטפורמה"
        bar.width = 20
        bar.height = 14
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=end)
        cats = Reference(ws, min_col=1, min_row=2, max_row=end)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "K2")


# ═══════════════════════════════════════════════════════════════
# SHEET 8: המלצות (Structured Recommendations)
# ═══════════════════════════════════════════════════════════════

def _build_recommendations_sheet(wb: Workbook, recommendations: list[dict]):
    """Structured recommendations: Campaign → Ad Set → Ad → Recommendation."""
    ws = wb.create_sheet("המלצות")
    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value="המלצות לשיפור ביצועים").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"תאריך: {date.today().strftime('%d/%m/%Y')}").font = Font(
        size=11, color="666666")

    headers = ["עדיפות", "קמפיין", "קבוצת מודעות", "מודעה", "סוג", "המלצה"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    type_map = {
        "budget": "תקציב",
        "kill": "לכבות",
        "scale": "להגדיל",
        "creative": "קריאייטיב",
        "audience": "קהל",
        "funnel": "פאנל",
        "frequency": "תדירות",
        "info": "מידע",
    }

    priority_map = {1: "🔴 דחוף", 2: "🟡 חשוב", 3: "🟢 הזדמנות", 4: "🔵 מידע", 5: "✅ תקין"}

    for i, rec in enumerate(recommendations, 5):
        priority = rec.get("priority", 5)
        ws.cell(row=i, column=1, value=priority_map.get(priority, ""))
        ws.cell(row=i, column=2, value=rec.get("campaign", ""))
        ws.cell(row=i, column=3, value=rec.get("adset", ""))
        ws.cell(row=i, column=4, value=rec.get("ad", ""))
        ws.cell(row=i, column=5, value=type_map.get(rec.get("type", ""), rec.get("type", "")))
        ws.cell(row=i, column=6, value=f"{rec.get('icon', '')} {rec.get('text', '')}")

        # Color by priority
        fill = PRIORITY_FILLS.get(priority, NEUTRAL_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).fill = fill
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).alignment = Alignment(
                horizontal="right" if c == 6 else "center",
                vertical="center", wrap_text=(c == 6))

        ws.row_dimensions[i].height = 35

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 65


# ═══════════════════════════════════════════════════════════════
# SHEET 9: השוואת תקופות (Period Comparison)
# ═══════════════════════════════════════════════════════════════

BETTER_FILL = PatternFill("solid", fgColor="C6EFCE")
WORSE_FILL = PatternFill("solid", fgColor="FFC7CE")
SAME_FILL = PatternFill("solid", fgColor="D9E2F3")


def _build_comparison_sheet(wb: Workbook, comparison: dict):
    ws = wb.create_sheet("השוואת תקופות")
    ws.sheet_view.rightToLeft = True

    row = 1

    for period_key, title in [("week", "שבוע נוכחי מול שבוע קודם"),
                               ("month", "חודש נוכחי מול חודש קודם")]:
        data = comparison.get(period_key)
        if not data:
            continue

        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 1
        ws.cell(row=row, column=1,
                value=f"{data['current_label']}  ←→  {data['previous_label']}").font = Font(
            size=11, color="666666")
        row += 1

        headers = ["מדד", "תקופה נוכחית", "תקופה קודמת", "שינוי %", "מגמה"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1

        for ch in data["changes"]:
            ws.cell(row=row, column=1, value=ch["label"])

            # Format values based on type
            fmt = ch["format"]
            cur = ch["current"]
            prev = ch["previous"]
            if fmt == "money":
                ws.cell(row=row, column=2, value=cur).number_format = MONEY_FMT2
                ws.cell(row=row, column=3, value=prev).number_format = MONEY_FMT2
            elif fmt == "decimal":
                ws.cell(row=row, column=2, value=round(cur, 2))
                ws.cell(row=row, column=3, value=round(prev, 2))
            elif fmt == "pct":
                ws.cell(row=row, column=2, value=round(cur, 2))
                ws.cell(row=row, column=3, value=round(prev, 2))
            else:
                ws.cell(row=row, column=2, value=cur)
                ws.cell(row=row, column=3, value=prev)

            pct = ch["pct_change"]
            arrow = "▲" if pct > 0 else ("▼" if pct < 0 else "—")
            ws.cell(row=row, column=4, value=f"{pct:+.1f}%")
            direction = ch["direction"]
            dir_text = {"better": "✅ שיפור", "worse": "⚠️ ירידה", "same": "➡️ יציב"}
            ws.cell(row=row, column=5, value=f"{arrow} {dir_text.get(direction, '')}")

            # Color row by direction
            fill = {"better": BETTER_FILL, "worse": WORSE_FILL, "same": SAME_FILL}.get(direction, SAME_FILL)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

        row += 2  # gap between sections

    _auto_width(ws, 5, 18)
    ws.column_dimensions["A"].width = 20


# ═══════════════════════════════════════════════════════════════
# SHEET 10: שחיקת קריאייטיב (Creative Fatigue)
# ═══════════════════════════════════════════════════════════════

FATIGUE_FILLS = {
    "critical": PatternFill("solid", fgColor="FFC7CE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
    "watch": PatternFill("solid", fgColor="D9E2F3"),
}


def _build_fatigue_sheet(wb: Workbook, fatigue_data: list[dict]):
    ws = wb.create_sheet("שחיקת קריאייטיב")
    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value="זיהוי שחיקת קריאייטיב").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="מודעות ש-CTR שלהן יורד ותדירות עולה — סימן שהקהל עייף מהקריאייטיב").font = Font(
        size=11, color="666666")

    if not fatigue_data:
        ws.cell(row=4, column=1, value="✅ לא זוהתה שחיקה במודעות הפעילות").font = Font(
            name="Arial", size=13, color="2F5496")
        return

    headers = ["חומרה", "קמפיין", "קבוצה", "מודעה",
               "CTR התחלה", "CTR עכשיו", "שינוי CTR%",
               "תדירות התחלה", "תדירות עכשיו", "שינוי תדירות%",
               "ימים", "הוצאה", "המלצה"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    sev_icons = {"critical": "🔥 קריטי", "warning": "😴 אזהרה", "watch": "👀 מעקב"}

    for i, f in enumerate(fatigue_data, 5):
        sev = f["severity"]
        ws.cell(row=i, column=1, value=sev_icons.get(sev, ""))
        ws.cell(row=i, column=2, value=f.get("campaign_name", ""))
        ws.cell(row=i, column=3, value=f.get("adset_name", ""))
        ws.cell(row=i, column=4, value=f.get("ad_name", ""))
        ws.cell(row=i, column=5, value=round(f["ctr_first"], 2))
        ws.cell(row=i, column=6, value=round(f["ctr_second"], 2))
        ws.cell(row=i, column=7, value=f"{f['ctr_change_pct']:+.0f}%")
        ws.cell(row=i, column=8, value=round(f["freq_first"], 1))
        ws.cell(row=i, column=9, value=round(f["freq_second"], 1))
        ws.cell(row=i, column=10, value=f"{f['freq_change_pct']:+.0f}%")
        ws.cell(row=i, column=11, value=f["days_running"])
        ws.cell(row=i, column=12, value=f["total_spend"])
        ws.cell(row=i, column=13, value=f["recommendation"])

        fill = FATIGUE_FILLS.get(sev, SAME_FILL)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=(c == 13))
        ws.cell(row=i, column=12).number_format = MONEY_FMT2
        ws.row_dimensions[i].height = 30

    _auto_width(ws, len(headers), 12)
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["M"].width = 40

    # ── CTR trend chart for top fatigued ads ──
    chart_row = 5 + len(fatigue_data) + 2
    for idx, f in enumerate(fatigue_data[:3]):  # Top 3
        daily = f.get("daily_ctr", [])
        if len(daily) < 3:
            continue

        ws.cell(row=chart_row, column=1,
                value=f"📉 {f['ad_name']}").font = SUBTITLE_FONT
        chart_row += 1

        # Write daily CTR data for chart
        ws.cell(row=chart_row, column=1, value="תאריך")
        ws.cell(row=chart_row, column=2, value="CTR%")
        ws.cell(row=chart_row, column=3, value="תדירות")
        data_start = chart_row + 1
        for j, d in enumerate(daily):
            ws.cell(row=data_start + j, column=1, value=d["date"])
            ws.cell(row=data_start + j, column=2, value=round(d["ctr"], 2))
            ws.cell(row=data_start + j, column=3, value=round(d["frequency"], 1))
        data_end = data_start + len(daily) - 1

        chart = LineChart()
        chart.title = f"CTR + תדירות — {f['ad_name'][:30]}"
        chart.width = 26
        chart.height = 12
        ctr_ref = Reference(ws, min_col=2, min_row=chart_row, max_row=data_end)
        freq_ref = Reference(ws, min_col=3, min_row=chart_row, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(ctr_ref, titles_from_data=True)
        chart.add_data(freq_ref, titles_from_data=True)
        chart.set_categories(cats)
        if chart.series:
            chart.series[0].graphicalProperties.line.width = 25000
        ws.add_chart(chart, f"E{chart_row}")

        chart_row = data_end + 2


# ═══════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def write_report(campaign_analysis: dict, ad_analysis: dict,
                 trend_analysis: dict, demo_analysis: dict,
                 placement_analysis: dict, recommendations: list[dict],
                 period_comparison: dict | None = None,
                 fatigue_data: list[dict] | None = None) -> Path:
    """Generate the full Excel report and return the file path."""
    output_dir = Path(settings.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"ads_report_{date.today().strftime('%Y%m%d')}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, campaign_analysis, trend_analysis)
    _build_campaigns_sheet(wb, campaign_analysis)
    _build_ads_sheet(wb, ad_analysis)
    _build_daily_sheet(wb, trend_analysis)
    _build_weekly_sheet(wb, trend_analysis)
    _build_monthly_sheet(wb, trend_analysis)
    if period_comparison:
        _build_comparison_sheet(wb, period_comparison)
    _build_demographics_sheet(wb, demo_analysis)
    _build_placements_sheet(wb, placement_analysis)
    if fatigue_data is not None:
        _build_fatigue_sheet(wb, fatigue_data)
    _build_recommendations_sheet(wb, recommendations)

    wb.save(output_path)
    logger.info("Report saved: %s", output_path)
    return output_path
