"""
WhatsApp notifier — sends ONE unified daily summary.
Yesterday's performance + P&L profit + 3 top recommendations.
Clean Hebrew RTL format.
"""

import logging
from datetime import date, timedelta
from pathlib import Path

import requests

from . import settings

logger = logging.getLogger(__name__)

API_URL = "https://graph.facebook.com/v22.0/{phone_id}/messages"

# Path to the P&L Excel file from monday-pnl
PNL_EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "monday-pnl" / "P&L.xlsx"


def _read_pnl_data(target_date: date | None = None) -> dict:
    """Read P&L summary for a specific date from the monday-pnl Excel file."""
    if target_date is None:
        target_date = date.today() - timedelta(days=1)

    result = {
        "total_revenue": 0,
        "total_profit": 0,
        "total_vat": 0,
        "total_cogs": 0,
        "total_shipping": 0,
        "total_payment_fee": 0,
        "total_orders": 0,
        "avg_margin": 0,
        "date": target_date.strftime("%d/%m/%Y"),
    }

    if not PNL_EXCEL_PATH.exists():
        logger.warning("P&L Excel not found at %s", PNL_EXCEL_PATH)
        return result

    try:
        from openpyxl import load_workbook
        wb = load_workbook(PNL_EXCEL_PATH, read_only=True, data_only=True)

        # Try to read from the calculation sheet
        sheet_name = None
        for name in wb.sheetnames:
            if name in ("חישוב", "Calc", "calc"):
                sheet_name = name
                break

        if not sheet_name:
            for name in wb.sheetnames:
                if name not in ("הגדרות", "Settings"):
                    sheet_name = name
                    break

        if not sheet_name:
            logger.warning("No calculation sheet found in P&L Excel")
            wb.close()
            return result

        ws = wb[sheet_name]

        # Find column indices from header row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                val = str(cell.value).strip()
                headers[val] = col_idx

        # Map Hebrew/English column names
        col_map = {
            "order_price": ["מחיר הזמנה", "order_price", "מחיר מכירה", "sell_price"],
            "profit": ["רווח", "profit"],
            "vat": ["מע״מ", "vat"],
            "cogs": ["עלות סחורה", "cogs"],
            "shipping": ["משלוח", "shipping"],
            "payment_fee": ["עמלת סליקה", "payment_fee"],
            "order_date": ["תאריך יצירה", "created_at", "תאריך הזמנה", "order_date"],
        }

        col_indices = {}
        for key, possible_names in col_map.items():
            for name in possible_names:
                if name in headers:
                    col_indices[key] = headers[name]
                    break

        if "profit" not in col_indices or "order_date" not in col_indices:
            logger.warning("Missing profit or order_date column in P&L Excel")
            wb.close()
            return result

        # Read rows and filter by target date
        total_revenue = 0
        total_profit = 0
        total_vat = 0
        total_cogs = 0
        total_shipping = 0
        total_payment_fee = 0
        total_orders = 0

        for row in ws.iter_rows(min_row=2, values_only=False):
            try:
                # Get order date
                date_val = row[col_indices["order_date"] - 1].value
                if date_val is None:
                    continue

                # Parse date — could be datetime, date, or string
                order_date = None
                if hasattr(date_val, 'date'):
                    order_date = date_val.date()
                elif isinstance(date_val, date):
                    order_date = date_val
                elif isinstance(date_val, str):
                    # Try common formats (including datetime strings with time component)
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
                        try:
                            from datetime import datetime
                            order_date = datetime.strptime(date_val.strip(), fmt).date()
                            break
                        except ValueError:
                            continue

                if order_date != target_date:
                    continue

                # This order is from the target date — sum it up
                order_price = 0
                profit = 0
                vat = 0
                cogs = 0
                shipping = 0
                payment_fee = 0

                if "order_price" in col_indices:
                    val = row[col_indices["order_price"] - 1].value
                    order_price = float(val) if val else 0

                if "profit" in col_indices:
                    val = row[col_indices["profit"] - 1].value
                    profit = float(val) if val else 0

                if "vat" in col_indices:
                    val = row[col_indices["vat"] - 1].value
                    vat = float(val) if val else 0

                if "cogs" in col_indices:
                    val = row[col_indices["cogs"] - 1].value
                    cogs = float(val) if val else 0

                if "shipping" in col_indices:
                    val = row[col_indices["shipping"] - 1].value
                    shipping = float(val) if val else 0

                if "payment_fee" in col_indices:
                    val = row[col_indices["payment_fee"] - 1].value
                    payment_fee = float(val) if val else 0

                if order_price > 0:
                    total_revenue += order_price
                    total_profit += profit
                    total_vat += vat
                    total_cogs += cogs
                    total_shipping += shipping
                    total_payment_fee += payment_fee
                    total_orders += 1

            except (ValueError, TypeError, IndexError):
                continue

        wb.close()

        result["total_revenue"] = round(total_revenue, 2)
        result["total_profit"] = round(total_profit, 2)
        result["total_vat"] = round(total_vat, 2)
        result["total_cogs"] = round(total_cogs, 2)
        result["total_shipping"] = round(total_shipping, 2)
        result["total_payment_fee"] = round(total_payment_fee, 2)
        result["total_orders"] = total_orders
        result["avg_margin"] = round(total_profit / total_revenue * 100, 1) if total_revenue > 0 else 0

        logger.info("P&L for %s: %d orders, revenue=%.0f, profit=%.0f, margin=%.1f%%",
                     target_date.strftime("%d/%m"), total_orders, total_revenue, total_profit, result["avg_margin"])

    except Exception:
        logger.exception("Failed to read P&L Excel")

    return result


def _pick_top_actions(recommendations: list[dict]) -> list[str]:
    """
    Pick the 3 most actionable recommendations — one per category.
    Returns short Hebrew action strings, no repetition.
    """
    chosen: list[str] = []
    seen_types: set[str] = set()

    # Priority order: kill → funnel → scale → frequency/fatigue → trend/audience
    type_order = ["kill", "funnel", "scale", "frequency", "fatigue", "creative", "budget", "audience", "info"]

    sorted_recs = sorted(recommendations, key=lambda r: (r["priority"], type_order.index(r["type"]) if r["type"] in type_order else 99))

    for rec in sorted_recs:
        if len(chosen) >= 3:
            break
        rtype = rec["type"]

        # Allow at most one per type, except scale (allow top 1 only)
        if rtype in seen_types:
            continue

        icon = rec.get("icon", "")
        text = rec.get("text", "")
        ad = rec.get("ad", "")
        adset = rec.get("adset", "")
        campaign = rec.get("campaign", "")

        # Build a short label — prefer ad name, then adset, then campaign
        label = ad or adset or (campaign if campaign != "כללי" else "")
        # Truncate long names
        if label and len(label) > 28:
            label = label[:26] + "…"

        # Rephrase into action-first, single-line format
        if rtype == "kill":
            chosen.append(f"{icon} *{label}* — הוצאה ₪{_extract_spend(text)}, 0 תוצאות → *לכבות*")
        elif rtype == "funnel":
            chosen.append(f"{icon} {text}")
        elif rtype == "scale":
            chosen.append(f"{icon} *{label}* — {_extract_roas_phrase(text)} → *להגדיל תקציב*")
        elif rtype in ("frequency", "fatigue"):
            chosen.append(f"{icon} *{label or 'כללי'}* — {_shorten(text, 55)}")
        elif rtype == "creative":
            chosen.append(f"{icon} {_shorten(text, 65)}")
        elif rtype == "budget":
            chosen.append(f"{icon} {_shorten(text, 65)}")
        elif rtype == "audience":
            chosen.append(f"{icon} {text}")
        else:
            chosen.append(f"{icon} {_shorten(text, 65)}")

        seen_types.add(rtype)

    return chosen


def _extract_spend(text: str) -> str:
    """Pull the ₪ amount from a recommendation text."""
    import re
    m = re.search(r"₪([\d,]+)", text)
    return m.group(1) if m else "?"


def _extract_roas_phrase(text: str) -> str:
    """Extract a short ROAS+purchases phrase."""
    import re
    roas_m = re.search(r"ROAS\s+([\d.]+)", text)
    purch_m = re.search(r"(\d+)\s+רכישות", text)
    parts = []
    if roas_m:
        parts.append(f"ROAS {roas_m.group(1)}")
    if purch_m:
        parts.append(f"{purch_m.group(1)} רכישות")
    return ", ".join(parts) if parts else text[:40]


def _shorten(text: str, max_len: int) -> str:
    return text if len(text) <= max_len else text[:max_len - 1] + "…"


def _build_daily_summary(
    campaign_data: dict,
    trend_data: dict,
    recommendations: list[dict],
    drive_link: str | None = None,
    hourly_data: list[dict] | None = None,
    video_data: list[dict] | None = None,
) -> str:
    """Build a clean, high-quality Hebrew daily WhatsApp summary.
    Sources:
    - Meta Ads API: spend, impressions, clicks, ATC, IC, CPC, CTR, hourly, video
    - WooCommerce: orders, revenue, visitors, CVR
    - ROAS = WooCommerce revenue / Meta spend
    """
    yesterday = date.today() - timedelta(days=1)

    # ── Meta: extract yesterday from daily breakdown ──
    daily_rows = trend_data.get("daily", [])
    yesterday_str = yesterday.isoformat()
    yesterday_data = next((d for d in daily_rows if d.get("date") == yesterday_str), None)

    if yesterday_data:
        spend       = yesterday_data.get("spend", 0)
        impressions = yesterday_data.get("impressions", 0)
        clicks      = yesterday_data.get("clicks", 0)
        atc         = yesterday_data.get("add_to_cart", 0)
        ic          = yesterday_data.get("initiate_checkout", 0)
        cpc         = yesterday_data.get("cpc", 0)
        ctr         = yesterday_data.get("ctr", 0)
    else:
        spend       = campaign_data.get("total_spend", 0)
        impressions = campaign_data.get("total_impressions", 0)
        clicks      = campaign_data.get("total_clicks", 0)
        atc         = campaign_data.get("total_atc", 0)
        ic          = campaign_data.get("total_ic", 0)
        cpc         = spend / clicks if clicks > 0 else 0
        ctr         = clicks / impressions * 100 if impressions > 0 else 0

    # ── Same weekday last week (from existing daily data — no extra API call) ──
    last_week = yesterday - timedelta(days=7)
    last_week_str = last_week.isoformat()
    lw = next((d for d in daily_rows if d.get("date") == last_week_str), None)
    lw_spend   = lw.get("spend", 0) if lw else 0
    lw_roas    = lw.get("roas", 0) if lw else 0

    def _delta(cur: float, prev: float) -> str:
        """Return a short ▲/▼ delta string."""
        if prev == 0:
            return ""
        pct = (cur - prev) / prev * 100
        arrow = "▲" if pct >= 0 else "▼"
        return f" {arrow}{abs(pct):.0f}%"

    # ── WooCommerce ──
    from .woo_client import get_daily_stats
    woo = get_daily_stats(target_date=yesterday)
    woo_lw  = get_daily_stats(target_date=last_week)
    woo_orders  = woo["orders"]
    woo_revenue = woo["revenue"]
    visitors    = woo["visitors"]
    cvr         = woo["conversion_rate"]
    roas        = woo_revenue / spend if spend > 0 else 0

    lw_woo_revenue = woo_lw["revenue"]
    lw_woo_orders  = woo_lw["orders"]

    # ── Funnel conversion rates ──
    atc_to_ic  = round(ic / atc * 100) if atc > 0 else 0
    ic_to_pur  = round(woo_orders / ic * 100) if ic > 0 else 0

    # ── Header ──
    lines = [
        f"*OneZone Jersey | {yesterday.strftime('%d/%m/%Y')}*",
        "",
    ]

    # ── Meta Ads block ──
    spend_delta = _delta(spend, lw_spend)
    lines += [
        "*📊 Meta Ads*",
        f"💸 הוצאה: *₪{spend:,.0f}*{spend_delta}",
        f"👁 חשיפות: {impressions:,}  |  CTR: {ctr:.2f}%",
        f"🛒 ATC: {atc}  |  IC: {ic}",
        f"💵 CPC: ₪{cpc:.2f}",
    ]

    # ── Funnel block ──
    if atc > 0:
        lines += [
            "",
            "*🔽 פאנל*",
            f"ATC → IC: {atc_to_ic}%  |  IC → רכישה: {ic_to_pur}%",
        ]

    # ── Site block ──
    lines += ["", "*🛍 אתר*"]
    if visitors > 0:
        lines.append(f"👤 מבקרים: {visitors:,}")
    if woo_orders > 0:
        orders_delta  = _delta(woo_orders, lw_woo_orders)
        revenue_delta = _delta(woo_revenue, lw_woo_revenue)
        roas_delta    = _delta(roas, lw_woo_revenue / lw_spend if lw_spend > 0 else 0)
        lines += [
            f"📦 הזמנות: *{woo_orders}*{orders_delta}  |  המרה: {cvr:.2f}%",
            f"💰 הכנסות: *₪{woo_revenue:,.0f}*{revenue_delta}",
            f"📈 ROAS: *{roas:.2f}x*{roas_delta}",
        ]
    else:
        lines.append(f"📦 אין הזמנות ב-{yesterday.strftime('%d/%m')}")

    # ── Hourly peaks block ──
    if hourly_data:
        # rank hours by ATC + purchases combined
        scored = sorted(
            [h for h in hourly_data if h["add_to_cart"] + h["purchases"] > 0],
            key=lambda h: h["add_to_cart"] + h["purchases"] * 3,
            reverse=True,
        )
        if scored:
            lines += ["", "*🕐 שעות שיא*"]
            for h in scored[:3]:
                hr = h["hour"]
                label = f"{hr:02d}:00–{hr+1:02d}:00"
                parts = []
                if h["add_to_cart"]:
                    parts.append(f"ATC {h['add_to_cart']}")
                if h["purchases"]:
                    parts.append(f"רכישות {h['purchases']}")
                lines.append(f"🕐 {label}: {' | '.join(parts)}")

    # ── Video retention block ──
    if video_data:
        # Show top 2 video ads by plays, sorted by p50 retention
        top_videos = sorted(
            [v for v in video_data if v["plays"] > 50],
            key=lambda v: v["p50_pct"],
            reverse=True,
        )[:2]
        if top_videos:
            lines += ["", "*📹 וידאו*"]
            for v in top_videos:
                name = v["ad_name"]
                if len(name) > 25:
                    name = name[:23] + "…"
                avg = f"{v['avg_watch_sec']}שנ" if v["avg_watch_sec"] else ""
                retention = f"25%▶{v['p25_pct']}%  50%▶{v['p50_pct']}%  100%▶{v['p100_pct']}%"
                lines.append(f"🎬 *{name}*")
                lines.append(f"   {retention}  {avg}")

    # ── Actions block ──
    actions = _pick_top_actions(recommendations)
    if actions:
        lines += ["", "*🎯 פעולות*"]
        for a in actions:
            lines.append(a)

    # ── Drive link ──
    if drive_link:
        lines += ["", f"📎 {drive_link}"]

    return "\n".join(lines)


def _send_via_template(
    yesterday: date,
    spend: float,
    purchases: int,
    roas: float,
    pnl_orders: int,
    net_profit: float,
) -> bool:
    """Send the daily report using the approved 'onezone_daily_report' template."""
    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID)
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "template",
        "template": {
            "name": "onezone_daily_report",
            "language": {"code": "he"},
            "components": [
                {
                    "type": "body",
                    "parameters": [
                        {"type": "text", "text": yesterday.strftime("%d/%m/%Y")},
                        {"type": "text", "text": f"{spend:,.0f}"},
                        {"type": "text", "text": str(purchases)},
                        {"type": "text", "text": f"{roas:.2f}"},
                        {"type": "text", "text": str(pnl_orders)},
                        {"type": "text", "text": f"{net_profit:,.0f}"},
                    ],
                }
            ],
        },
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        msg_id = resp.json().get("messages", [{}])[0].get("id", "unknown")
        logger.info("Template report sent (ID: %s)", msg_id)
        return True
    except requests.exceptions.HTTPError as e:
        logger.error("Template send failed: %s — %s", e, resp.text)
        return False
    except Exception:
        logger.exception("Failed to send template report")
        return False


def _upload_media(file_path: str) -> str | None:
    """Upload a file to WhatsApp media API. Returns media_id or None."""
    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID).replace("/messages", "/media")
    headers = {"Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}"}
    try:
        with open(file_path, "rb") as f:
            resp = requests.post(
                url,
                headers=headers,
                files={"file": (file_path.split("\\")[-1].split("/")[-1], f, "application/pdf")},
                data={"messaging_product": "whatsapp", "type": "document"},
                timeout=60,
            )
        resp.raise_for_status()
        media_id = resp.json().get("id")
        logger.info("Media uploaded — ID: %s", media_id)
        return media_id
    except Exception:
        logger.exception("Failed to upload media")
        return None


def _send_via_document(pdf_path: str, filename: str, caption: str = "") -> bool:
    """Upload PDF to WhatsApp and send as document message with optional caption."""
    media_id = _upload_media(pdf_path)
    if not media_id:
        logger.warning("Media upload failed — falling back to text message")
        return False

    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID)
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    doc_obj: dict = {"id": media_id, "filename": filename}
    if caption:
        doc_obj["caption"] = caption
    payload = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "document",
        "document": doc_obj,
    }
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        msg_id = resp.json().get("messages", [{}])[0].get("id", "unknown")
        logger.info("PDF document sent (ID: %s)", msg_id)
        return True
    except requests.exceptions.HTTPError as e:
        logger.error("Document send failed: %s — %s", e, resp.text)
        return False
    except Exception:
        logger.exception("Failed to send document")
        return False


def _send_via_text(message: str) -> bool:
    """Send full text report directly (works within 24h conversation window)."""
    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID)
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    text_payload = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "text",
        "text": {"body": message},
    }
    try:
        resp = requests.post(url, json=text_payload, headers=headers, timeout=30)
        resp.raise_for_status()
        msg_id = resp.json().get("messages", [{}])[0].get("id", "unknown")
        logger.info("Text report sent (ID: %s)", msg_id)
        return True
    except requests.exceptions.HTTPError as e:
        logger.error("Text report failed: %s — %s", e, resp.text)
        return False
    except Exception:
        logger.exception("Failed to send text report")
        return False


def send_error_alert(pipeline: str, error: str) -> bool:
    """
    Send a WhatsApp alert when a pipeline run fails.
    Uses hello_world to open the conversation window, then sends the error.
    """
    if not settings.WHATSAPP_ACCESS_TOKEN or not settings.WHATSAPP_PHONE_NUMBER_ID:
        logger.warning("WhatsApp credentials not configured, cannot send error alert")
        return False
    if not settings.WHATSAPP_TO:
        return False

    import time

    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID)
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }

    # Truncate long errors so message stays readable
    error_short = error[:300] + "..." if len(error) > 300 else error

    message = (
        f"⚠️ *שגיאה ב-{pipeline}*\n"
        f"📅 {date.today().strftime('%d/%m/%Y %H:%M')}\n\n"
        f"הדוח היומי *לא נשלח* בגלל שגיאה:\n"
        f"```{error_short}```\n\n"
        f"בדוק את הלוגים ב-logs/ads.log"
    )

    # Open 24h conversation window first
    opener = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "template",
        "template": {"name": "hello_world", "language": {"code": "en_US"}},
    }
    try:
        resp = requests.post(url, json=opener, headers=headers, timeout=30)
        resp.raise_for_status()
        time.sleep(3)
    except Exception:
        logger.exception("Error alert: hello_world opener failed")
        return False

    text_payload = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "text",
        "text": {"body": message},
    }
    try:
        resp = requests.post(url, json=text_payload, headers=headers, timeout=30)
        resp.raise_for_status()
        logger.info("Error alert sent for pipeline '%s'", pipeline)
        return True
    except Exception:
        logger.exception("Failed to send error alert")
        return False


def send_token_expiry_warning(days_left: int) -> bool:
    """Send a WhatsApp warning when the Facebook API token is about to expire."""
    if not settings.WHATSAPP_ACCESS_TOKEN or not settings.WHATSAPP_PHONE_NUMBER_ID:
        return False
    if not settings.WHATSAPP_TO:
        return False

    import time

    url = API_URL.format(phone_id=settings.WHATSAPP_PHONE_NUMBER_ID)
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }

    urgency = "🔴" if days_left <= 3 else "🟡"
    message = (
        f"{urgency} *טוקן פייסבוק פג בעוד {days_left} ימים*\n\n"
        f"הטוקן של Meta Ads API יפוג בקרוב.\n"
        f"אם לא תחדש אותו — הדוחות היומיים יפסיקו לעבוד.\n\n"
        f"*איך לחדש:*\n"
        f"1. כנס ל-business.facebook.com\n"
        f"2. System Users → Generate Token\n"
        f"3. עדכן את FB_ACCESS_TOKEN בקובץ .env"
    )

    opener = {
        "messaging_product": "whatsapp",
        "to": settings.WHATSAPP_TO,
        "type": "template",
        "template": {"name": "hello_world", "language": {"code": "en_US"}},
    }
    try:
        resp = requests.post(url, json=opener, headers=headers, timeout=30)
        resp.raise_for_status()
        time.sleep(3)
    except Exception:
        logger.exception("Token warning: opener failed")
        return False

    try:
        resp = requests.post(url, json={
            "messaging_product": "whatsapp",
            "to": settings.WHATSAPP_TO,
            "type": "text",
            "text": {"body": message},
        }, headers=headers, timeout=30)
        resp.raise_for_status()
        logger.info("Token expiry warning sent (%d days left)", days_left)
        return True
    except Exception:
        logger.exception("Failed to send token expiry warning")
        return False


def send_summary(
    campaign_data: dict,
    ad_data: dict,
    trend_data: dict,
    recommendations: list,
    drive_link: str | None = None,
    structured_recs: list[dict] | None = None,
    hourly_data: list[dict] | None = None,
    video_data: list[dict] | None = None,
) -> bool:
    """Send unified daily summary via WhatsApp. Returns True on success."""
    if not settings.WHATSAPP_ACCESS_TOKEN or not settings.WHATSAPP_PHONE_NUMBER_ID:
        logger.warning("WhatsApp credentials not configured, skipping notification")
        return False

    if not settings.WHATSAPP_TO:
        logger.warning("WhatsApp recipient not configured (WHATSAPP_TO), skipping")
        return False

    yesterday = date.today() - timedelta(days=1)

    # ── Extract yesterday's data ──
    daily_rows = trend_data.get("daily", [])
    yesterday_str = yesterday.isoformat()
    yesterday_data = None
    for d in daily_rows:
        if d.get("date") == yesterday_str:
            yesterday_data = d
            break

    if yesterday_data:
        spend = yesterday_data.get("spend", 0)
        purchases = yesterday_data.get("purchases", 0)
        roas = yesterday_data.get("roas", 0)
    else:
        spend = campaign_data.get("total_spend", 0)
        purchases = campaign_data.get("total_purchases", 0)
        roas = campaign_data.get("total_roas", 0)

    pnl = _read_pnl_data(target_date=yesterday)
    net_profit = pnl["total_profit"] - spend

    recs = structured_recs or []
    if not recs and recommendations:
        recs = [{"icon": "", "text": r, "campaign": "", "ad": ""} for r in recommendations]

    actions = _pick_top_actions(recs)

    # ── Try to generate & send PDF ──
    try:
        from .pdf_report import generate_daily_pdf
        from .woo_client import get_daily_stats

        woo      = get_daily_stats(target_date=yesterday)
        woo_lw   = get_daily_stats(target_date=yesterday - timedelta(days=7))

        daily_rows   = trend_data.get("daily", [])
        yesterday_data = next((d for d in daily_rows if d.get("date") == yesterday.isoformat()), None)
        lw_data        = next((d for d in daily_rows if d.get("date") == (yesterday - timedelta(days=7)).isoformat()), None)

        spend       = yesterday_data.get("spend", 0) if yesterday_data else campaign_data.get("total_spend", 0)
        impressions = yesterday_data.get("impressions", 0) if yesterday_data else 0
        clicks      = yesterday_data.get("clicks", 0) if yesterday_data else 0
        atc         = yesterday_data.get("add_to_cart", 0) if yesterday_data else 0
        ic          = yesterday_data.get("initiate_checkout", 0) if yesterday_data else 0
        cpc         = yesterday_data.get("cpc", 0) if yesterday_data else 0
        ctr         = yesterday_data.get("ctr", 0) if yesterday_data else 0

        meta_data = {"spend": spend, "impressions": impressions, "clicks": clicks,
                     "atc": atc, "ic": ic, "cpc": cpc, "ctr": ctr}
        meta_lw   = {"spend": lw_data.get("spend", 0)} if lw_data else {"spend": 0}
        funnel    = {
            "atc_to_ic":       round(ic / atc * 100) if atc > 0 else 0,
            "ic_to_purchase":  round(woo["orders"] / ic * 100) if ic > 0 else 0,
        }
        woo_data  = {"orders": woo["orders"], "revenue": woo["revenue"],
                     "visitors": woo["visitors"], "cvr": woo["conversion_rate"]}
        woo_lw_data = {"orders": woo_lw["orders"], "revenue": woo_lw["revenue"]}

        pdf_path = generate_daily_pdf(
            report_date=yesterday,
            meta=meta_data, meta_lw=meta_lw,
            woo=woo_data, woo_lw=woo_lw_data,
            funnel=funnel,
            hourly_data=hourly_data or [],
            video_data=video_data or [],
            actions=actions,
            drive_link=drive_link,
        )
        filename = f"OneZone_{yesterday.strftime('%d_%m_%Y')}.pdf"
        # Send title text first, then the PDF document
        _send_via_text(f"📋 *סיכום יומי — {yesterday.strftime('%d/%m/%Y')}*")
        if _send_via_document(str(pdf_path), filename):
            return True
        logger.warning("PDF send failed — falling back to text")
    except Exception:
        logger.exception("PDF generation/send failed — falling back to text")

    # ── Fallback: plain text ──
    message = _build_daily_summary(campaign_data, trend_data, recs, drive_link,
                                   hourly_data=hourly_data, video_data=video_data)
    return _send_via_text(message)
