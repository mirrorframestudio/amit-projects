"""
Excel report generator — produces a Hebrew RTL P&L workbook with 3 tabs:
1. הגדרות (Settings)
2. הזמנות (Orders)
3. סיכום שבועי (Weekly Summary)
"""
import logging
import os
import sys
from datetime import date
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .pnl_engine import OrderPnL, WeeklySummary
from .settings import SETTINGS

# ── Shared styles (single source of truth for all P&L Excel files) ──
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from pnl_shared.excel_styles import (  # noqa: E402
    HEADER_FONT,
    HEADER_FILL,
    HEADER_ALIGN,
    THIN_BORDER,
    TITLE_FONT,
    KPI_FONT as SUMMARY_LABEL_FONT,
    BODY_FONT as SUMMARY_VAL_FONT,
    WARNING_FILL,
    PROFIT_NEG_FONT as WARNING_FONT,
    CURRENCY_FMT,
    PERCENT_FMT,
    set_rtl as _set_rtl,
    auto_width as _auto_width,
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════
# TAB 1: הגדרות (Settings)
# ══════════════════════════════════════════════════════════════════════
def _build_settings_sheet(wb: Workbook, week_start: date, week_end: date):
    ws = wb.create_sheet("הגדרות")
    _set_rtl(ws)

    ws.merge_cells("A1:B1")
    ws["A1"] = "הגדרות דוח רווח והפסד"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="right")

    rows = [
        ("תאריך התחלה", str(week_start)),
        ("תאריך סיום", str(week_end)),
        ("שיעור מע״מ", f"{SETTINGS.VAT_RATE:.0%}"),
        ("עמלת סליקה (%)", f"{SETTINGS.EXTERNAL_PROCESSING_FEE_PERCENT:.1%}"),
        ("עמלת סליקה קבועה (₪)", f"₪{SETTINGS.EXTERNAL_PROCESSING_FEE_FIXED:.2f}"),
        ("עלות משלוח — נקודת איסוף (₪)", f"₪{SETTINGS.SHIPPING_COSTS['PICKUP_POINT']:.2f}"),
        ("עלות משלוח — משלוח עד הבית (₪)", f"₪{SETTINGS.SHIPPING_COSTS['HOME_DELIVERY']:.2f}"),
        ("שיטת הקצאת שיווק", "יחסי (לפי מכירות)" if SETTINGS.MARKETING_ALLOCATION == "proportional" else "שווה (לפי הזמנה)"),
        ("קובץ שיווק", SETTINGS.MARKETING_CSV_PATH),
        ("מטבע דיווח", "ILS (₪)"),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = SUMMARY_LABEL_FONT
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=2, value=value).font = SUMMARY_VAL_FONT
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="right")

    _auto_width(ws)


# ══════════════════════════════════════════════════════════════════════
# TAB 2: הזמנות (Orders)
# ══════════════════════════════════════════════════════════════════════
ORDER_HEADERS = [
    "תחילת שבוע",
    "סוף שבוע",
    "מזהה הזמנה",
    "מספר הזמנה",
    "תאריך הזמנה",
    "סטטוס תשלום",
    "סטטוס אספקה",
    "מטבע",
    "כותרת שיטת משלוח",
    "סוג משלוח",
    "סכום ביניים",
    "הנחות",
    "עלות משלוח (גבייה)",
    "מע״מ/מס",
    "סה״כ ששולם",
    "בסיס הכנסה",
    "מע״מ שהוסר",
    "מכירות נטו ללא מע״מ",
    "עמלת סליקה",
    "שיווק מוקצה",
    "עלות משלוח (עלות)",
    "עלות מוצר (COGS)",
    "רווח נקי",
    "אחוז רווח",
    "סכום החזר",
    "דגלים/הערות",
]

# Columns that are currency (0-indexed positions)
CURRENCY_COLS = {10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 24}
PERCENT_COLS = {23}


def _build_orders_sheet(wb: Workbook, pnl_rows: List[OrderPnL]):
    ws = wb.create_sheet("הזמנות")
    _set_rtl(ws)

    # Header row
    for col_idx, header in enumerate(ORDER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, r in enumerate(pnl_rows, start=2):
        data = [
            str(r.week_start),
            str(r.week_end),
            r.order_id,
            r.order_name,
            r.created_at[:10] if r.created_at else "",
            r.financial_status,
            r.fulfillment_status,
            r.currency,
            r.shipping_title,
            r.shipping_type,
            r.subtotal,
            r.discounts,
            r.shipping_revenue,
            r.tax_amount,
            r.total_paid,
            r.revenue_base,
            r.vat_removed,
            r.net_sales_ex_vat,
            r.processing_fee,
            r.marketing_allocated,
            r.shipping_cost,
            r.cogs,
            r.net_profit,
            r.net_margin,
            r.refund_amount,
            " | ".join(r.flags) if r.flags else "",
        ]
        for col_idx, val in enumerate(data):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if col_idx in CURRENCY_COLS:
                cell.number_format = CURRENCY_FMT
            elif col_idx in PERCENT_COLS:
                cell.number_format = PERCENT_FMT

        # Highlight flags column if non-empty
        flags_cell = ws.cell(row=row_idx, column=len(ORDER_HEADERS))
        if flags_cell.value:
            flags_cell.fill = WARNING_FILL
            flags_cell.font = WARNING_FONT

    # Freeze header row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ORDER_HEADERS))}{len(pnl_rows) + 1}"

    _auto_width(ws)


# ══════════════════════════════════════════════════════════════════════
# TAB 3: סיכום שבועי (Weekly Summary)
# ══════════════════════════════════════════════════════════════════════
def _build_summary_sheet(wb: Workbook, summary: WeeklySummary):
    ws = wb.create_sheet("סיכום שבועי")
    _set_rtl(ws)

    ws.merge_cells("A1:B1")
    ws["A1"] = "סיכום רווח והפסד שבועי"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="right")

    rows = [
        ("טווח תאריכים", f"{summary.week_start}  —  {summary.week_end}"),
        ("כמות הזמנות", summary.order_count),
        ("", ""),  # spacer
        ("סה״כ בסיס הכנסה", summary.total_revenue_base),
        ("סה״כ מע״מ שהוסר", summary.total_vat_removed),
        ("סה״כ מכירות נטו ללא מע״מ", summary.total_net_sales_ex_vat),
        ("", ""),
        ("סה״כ עמלת סליקה", summary.total_processing_fees),
        ("סה״כ שיווק (מהקובץ)", summary.total_marketing_from_csv),
        ("סה״כ שיווק מוקצה", summary.total_marketing_allocated),
        ("סה״כ עלות משלוח", summary.total_shipping_cost),
        ("סה״כ עלות מוצר (COGS)", summary.total_cogs),
        ("", ""),
        ("סה״כ רווח נקי", summary.total_net_profit),
        ("אחוז רווח נקי", summary.net_margin),
        ("", ""),
        ("── איכות נתונים ──", ""),
        ("הזמנות עם עלות מוצר חסרה", summary.missing_cost_count),
        ("הזמנות עם סוג משלוח לא מזוהה", summary.unknown_shipping_count),
    ]

    CURRENCY_ROWS = {3, 4, 5, 7, 8, 9, 10, 11, 13}  # 0-indexed data row
    PERCENT_ROWS = {14}

    for i, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = SUMMARY_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="right")

        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = SUMMARY_VAL_FONT
        val_cell.alignment = Alignment(horizontal="right")

        data_idx = i - 3
        if data_idx in CURRENCY_ROWS and isinstance(value, (int, float)):
            val_cell.number_format = CURRENCY_FMT
        elif data_idx in PERCENT_ROWS and isinstance(value, (int, float)):
            val_cell.number_format = PERCENT_FMT

    # Highlight net profit row
    np_row = 3 + 13  # row 16
    ws.cell(row=np_row, column=1).font = Font(name="Arial", bold=True, size=12, color="006100")
    ws.cell(row=np_row, column=2).font = Font(name="Arial", bold=True, size=12, color="006100")
    ws.cell(row=np_row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Warnings
    warn_row = 3 + len(rows) + 1
    if summary.warnings:
        ws.cell(row=warn_row, column=1, value="── אזהרות ──").font = Font(
            name="Arial", bold=True, size=11, color="9C0006"
        )
        for j, w in enumerate(summary.warnings):
            cell = ws.cell(row=warn_row + 1 + j, column=1, value=w)
            cell.font = WARNING_FONT
            cell.fill = WARNING_FILL

    # Marketing missing — prominent warning
    if summary.marketing_missing:
        r = warn_row + len(summary.warnings) + 2
        cell = ws.cell(row=r, column=1, value="⚠ חסרה שורת שיווק בקובץ — הוצאת שיווק = 0")
        cell.font = Font(name="Arial", bold=True, size=12, color="9C0006")
        cell.fill = WARNING_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


# ══════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════
def generate_report(
    pnl_rows: List[OrderPnL],
    summary: WeeklySummary,
    week_start: date,
    week_end: date,
    output_dir: str = "",
) -> str:
    """
    Generate the Excel workbook and return the output file path.
    """
    out = output_dir or SETTINGS.OUTPUT_DIR
    os.makedirs(out, exist_ok=True)

    filename = f"pnl_week_{week_start}_to_{week_end}.xlsx"
    filepath = os.path.join(out, filename)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _build_settings_sheet(wb, week_start, week_end)
    _build_orders_sheet(wb, pnl_rows)
    _build_summary_sheet(wb, summary)

    wb.save(filepath)
    logger.info(f"Report saved to {filepath}")
    return filepath
