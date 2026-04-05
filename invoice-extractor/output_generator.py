import re
from datetime import datetime
from email.utils import parsedate_to_datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
GREEN = "1E8449"
RED = "C0392B"


def _parse_email_date(raw: str) -> str:
    try:
        dt = parsedate_to_datetime(raw)
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return raw or ''


def _clean_sender(sender: str) -> str:
    m = re.match(r'^"?([^"<]+)"?\s*<', sender)
    return m.group(1).strip() if m else sender.split('<')[0].strip()


class OutputGenerator:
    def create_excel(self, invoices: list[dict], output_path: str = None) -> str:
        if not output_path:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'הוצאות_{ts}.xlsx'

        df = self._build_df(invoices)
        df.to_excel(output_path, index=False, sheet_name='הוצאות')

        wb = load_workbook(output_path)
        ws = wb['הוצאות']
        ws.sheet_view.rightToLeft = True
        self._style_header(ws)
        self._style_data(ws, df)
        self._add_total_row(ws, df)
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'
        wb.save(output_path)
        return output_path

    def _build_df(self, invoices: list[dict]) -> 'pd.DataFrame':
        rows = []
        for inv in invoices:
            rows.append({
                'תאריך מייל': _parse_email_date(inv.get('email_date', '')),
                'שולח': _clean_sender(inv.get('sender', '')),
                'נושא מייל': inv.get('email_subject', ''),
                'מספר חשבונית': inv.get('invoice_number', ''),
                'ספק': inv.get('supplier', ''),
                'תאריך חשבונית': inv.get('date', ''),
                'מטבע': inv.get('currency', 'ILS'),
                'סכום': inv.get('amount'),
                'קובץ חשבונית': inv.get('saved_file', inv.get('filename', '')),
                'הערות': inv.get('error', ''),
            })
        return pd.DataFrame(rows)

    # ------------------------------------------------------------------ #

    def _style_header(self, ws):
        header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[1].height = 28

    def _style_data(self, ws, df):
        amount_col_idx = list(df.columns).index('סכום') + 1
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx in range(2, ws.max_row + 1):
            is_even = (row_idx % 2 == 0)
            row_fill = PatternFill(
                start_color=LIGHT_BLUE if is_even else 'FFFFFF',
                end_color=LIGHT_BLUE if is_even else 'FFFFFF',
                fill_type='solid'
            )
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # Amount column: bold, number format
            amount_cell = ws.cell(row=row_idx, column=amount_col_idx)
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = Alignment(horizontal='center', vertical='center')

    def _add_total_row(self, ws, df):
        total_row = ws.max_row + 1
        amount_col = list(df.columns).index('סכום') + 1

        label_cell = ws.cell(row=total_row, column=amount_col - 1, value='סה"כ:')
        label_cell.font = Font(bold=True, size=12, name='Calibri')
        label_cell.alignment = Alignment(horizontal='right', vertical='center')

        total_cell = ws.cell(row=total_row, column=amount_col)
        col_letter = get_column_letter(amount_col)
        total_cell.value = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'
        total_cell.font = Font(bold=True, size=13, color=DARK_BLUE, name='Calibri')
        total_cell.number_format = '₪#,##0.00'
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[total_row].height = 30

    def _autofit_columns(self, ws):
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
                default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
