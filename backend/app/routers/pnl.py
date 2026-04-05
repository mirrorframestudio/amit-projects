"""
P&L (Profit & Loss) aggregation endpoint.

Queries multiple tables and aggregates into a P&L summary by period.
"""

import io
import os
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

from .. import schemas
from ..database import get_db

router = APIRouter(prefix="/pnl", tags=["pnl"])

# Platform fee defaults (Shopify Payments standard rate)
PLATFORM_FEE_RATE = float(os.getenv("PLATFORM_FEE_RATE", "0.026"))
PLATFORM_FEE_FIXED = float(os.getenv("PLATFORM_FEE_FIXED", "0.30"))


@router.get("/summary", response_model=schemas.PnLSummary)
def get_pnl_summary(
    start_date: date = Query(..., description="Start date (inclusive)"),
    end_date: date = Query(..., description="End date (inclusive)"),
    resolution: str = Query("monthly", description="daily, weekly, or monthly"),
    db: Session = Depends(get_db),
):
    # Determine date_trunc interval
    if resolution == "daily":
        trunc = "day"
        fmt = "YYYY-MM-DD"
    elif resolution == "weekly":
        trunc = "week"
        fmt = "IYYY-\"W\"IW"
    else:
        trunc = "month"
        fmt = "YYYY-MM"

    # Validate trunc to prevent SQL injection (it's interpolated directly)
    assert trunc in ("day", "week", "month")

    query = text(f"""
        WITH periods AS (
            SELECT generate_series(
                date_trunc('{trunc}', CAST(:start_date AS date)),
                date_trunc('{trunc}', CAST(:end_date AS date)),
                CAST('1 {trunc}' AS interval)
            )::date AS period
        ),

        -- Revenue from Shopify orders
        revenue AS (
            SELECT
                date_trunc('{trunc}', created_at_shopify)::date AS period,
                COALESCE(SUM(subtotal_price), 0) AS revenue,
                COALESCE(SUM(total_refunded), 0) AS refunds,
                COALESCE(SUM(total_shipping), 0) AS shipping,
                COUNT(*) AS order_count
            FROM shopify_orders
            WHERE created_at_shopify >= CAST(:start_date AS date)
              AND created_at_shopify < CAST(:end_date AS date) + 1
            GROUP BY 1
        ),

        -- COGS: join order items with product costs by SKU + date range
        cogs AS (
            SELECT
                date_trunc('{trunc}', so.created_at_shopify)::date AS period,
                COALESCE(SUM(
                    soi.quantity * COALESCE(pc.cost_price, 0)
                ), 0) AS cogs
            FROM shopify_order_items soi
            JOIN shopify_orders so ON so.id = soi.order_id
            LEFT JOIN product_costs pc
                ON pc.sku = soi.sku
                AND so.created_at_shopify::date >= pc.effective_date
                AND (pc.end_date IS NULL OR so.created_at_shopify::date <= pc.end_date)
            WHERE so.created_at_shopify >= CAST(:start_date AS date)
              AND so.created_at_shopify < CAST(:end_date AS date) + 1
            GROUP BY 1
        ),

        -- Ad spend from manual_expenses (category = 'ad_spend')
        ad_spend AS (
            SELECT
                date_trunc('{trunc}', me.start_date)::date AS period,
                COALESCE(SUM(
                    CASE me.frequency
                        WHEN 'one_time' THEN me.amount
                        WHEN 'daily' THEN me.amount *
                            CASE '{trunc}'
                                WHEN 'day' THEN 1
                                WHEN 'week' THEN 7
                                WHEN 'month' THEN 30
                            END
                        WHEN 'monthly' THEN me.amount
                        WHEN 'annual' THEN me.amount / 12.0
                    END
                ), 0) AS ad_spend
            FROM manual_expenses me
            WHERE me.category = 'ad_spend'
              AND me.start_date >= CAST(:start_date AS date)
              AND me.start_date <= CAST(:end_date AS date)
            GROUP BY 1
        ),

        -- Fixed expenses (all categories except ad_spend)
        fixed AS (
            SELECT
                p.period,
                COALESCE(SUM(
                    CASE me.frequency
                        WHEN 'one_time' THEN me.amount
                        WHEN 'daily' THEN me.amount *
                            CASE '{trunc}'
                                WHEN 'day' THEN 1
                                WHEN 'week' THEN 7
                                WHEN 'month' THEN 30
                            END
                        WHEN 'monthly' THEN me.amount
                        WHEN 'annual' THEN me.amount / 12.0
                    END
                ), 0) AS fixed_expenses
            FROM periods p
            LEFT JOIN manual_expenses me
                ON me.category != 'ad_spend'
                AND me.start_date <= (p.period + CAST('1 {trunc}' AS interval) - interval '1 day')::date
                AND (me.end_date IS NULL OR me.end_date >= p.period)
            GROUP BY 1
        )

        SELECT
            to_char(p.period, :fmt) AS period_label,
            COALESCE(r.revenue, 0) AS revenue,
            COALESCE(r.refunds, 0) AS refunds,
            COALESCE(r.shipping, 0) AS shipping,
            COALESCE(r.order_count, 0) AS order_count,
            COALESCE(c.cogs, 0) AS cogs,
            COALESCE(a.ad_spend, 0) AS ad_spend,
            COALESCE(f.fixed_expenses, 0) AS fixed_expenses
        FROM periods p
        LEFT JOIN revenue r ON r.period = p.period
        LEFT JOIN cogs c ON c.period = p.period
        LEFT JOIN ad_spend a ON a.period = p.period
        LEFT JOIN fixed f ON f.period = p.period
        ORDER BY p.period
    """)

    rows = db.execute(query, {
        "fmt": fmt,
        "start_date": str(start_date),
        "end_date": str(end_date),
    }).fetchall()

    # Build P&L line items
    pnl_rows = []
    totals = {
        "revenue": 0, "refunds": 0, "net_revenue": 0, "cogs": 0,
        "gross_profit": 0, "ad_spend": 0, "shipping": 0,
        "platform_fees": 0, "fixed_expenses": 0, "total_expenses": 0,
        "net_profit": 0,
    }

    for row in rows:
        revenue = float(row.revenue)
        refunds = float(row.refunds)
        net_revenue = revenue - refunds
        cogs = float(row.cogs)
        gross_profit = net_revenue - cogs
        ad_spend = float(row.ad_spend)
        shipping = float(row.shipping)
        order_count = int(row.order_count)
        platform_fees = round(revenue * PLATFORM_FEE_RATE + PLATFORM_FEE_FIXED * order_count, 2)
        fixed_expenses = float(row.fixed_expenses)
        total_expenses = ad_spend + shipping + platform_fees + fixed_expenses
        net_profit = gross_profit - total_expenses
        gross_margin = (gross_profit / net_revenue * 100) if net_revenue else 0
        net_margin = (net_profit / net_revenue * 100) if net_revenue else 0

        item = schemas.PnLLineItem(
            period=row.period_label,
            revenue=revenue,
            refunds=refunds,
            net_revenue=net_revenue,
            cogs=cogs,
            gross_profit=gross_profit,
            gross_margin=round(gross_margin, 1),
            ad_spend=ad_spend,
            shipping=shipping,
            platform_fees=platform_fees,
            fixed_expenses=fixed_expenses,
            total_expenses=total_expenses,
            net_profit=net_profit,
            net_margin=round(net_margin, 1),
        )
        pnl_rows.append(item)

        # Accumulate totals
        for key in totals:
            totals[key] += getattr(item, key, 0)

    # Calculate total margins
    total_gross_margin = (totals["gross_profit"] / totals["net_revenue"] * 100) if totals["net_revenue"] else 0
    total_net_margin = (totals["net_profit"] / totals["net_revenue"] * 100) if totals["net_revenue"] else 0

    totals_item = schemas.PnLLineItem(
        period="Total",
        revenue=round(totals["revenue"], 2),
        refunds=round(totals["refunds"], 2),
        net_revenue=round(totals["net_revenue"], 2),
        cogs=round(totals["cogs"], 2),
        gross_profit=round(totals["gross_profit"], 2),
        gross_margin=round(total_gross_margin, 1),
        ad_spend=round(totals["ad_spend"], 2),
        shipping=round(totals["shipping"], 2),
        platform_fees=round(totals["platform_fees"], 2),
        fixed_expenses=round(totals["fixed_expenses"], 2),
        total_expenses=round(totals["total_expenses"], 2),
        net_profit=round(totals["net_profit"], 2),
        net_margin=round(total_net_margin, 1),
    )

    return schemas.PnLSummary(
        start_date=start_date,
        end_date=end_date,
        resolution=resolution,
        currency="ILS",
        rows=pnl_rows,
        totals=totals_item,
    )


@router.get("/export")
def export_pnl_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    resolution: str = Query("monthly"),
    db: Session = Depends(get_db),
):
    """Export P&L summary as a styled Excel file."""
    summary = get_pnl_summary(start_date, end_date, resolution, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    # -- Styles --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    total_fill = PatternFill("solid", fgColor="e8e8e8")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )
    money_fmt = '#,##0.00 ₪'
    pct_fmt = '0.0"%"'

    # Row labels for the P&L lines
    line_labels = [
        ("Revenue", "revenue", money_fmt),
        ("Refunds", "refunds", money_fmt),
        ("Net Revenue", "net_revenue", money_fmt),
        ("COGS", "cogs", money_fmt),
        ("Gross Profit", "gross_profit", money_fmt),
        ("Gross Margin", "gross_margin", pct_fmt),
        ("Ad Spend", "ad_spend", money_fmt),
        ("Shipping", "shipping", money_fmt),
        ("Platform Fees", "platform_fees", money_fmt),
        ("Fixed Expenses", "fixed_expenses", money_fmt),
        ("Total Expenses", "total_expenses", money_fmt),
        ("Net Profit", "net_profit", money_fmt),
        ("Net Margin", "net_margin", pct_fmt),
    ]

    # -- Title row --
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary.rows) + 2)
    title_cell = ws.cell(row=1, column=1, value=f"P&L Report  |  {start_date} → {end_date}  |  {resolution}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # -- Header row (row 3) --
    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    for col_idx, item in enumerate(summary.rows, start=2):
        ws.cell(row=header_row, column=col_idx, value=item.period)
    ws.cell(row=header_row, column=len(summary.rows) + 2, value="Total")

    # Style header
    for col in range(1, len(summary.rows) + 3):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # -- Data rows --
    for row_idx, (label, field, fmt) in enumerate(line_labels, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=label in ("Net Revenue", "Gross Profit", "Net Profit"))

        for col_idx, item in enumerate(summary.rows, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=getattr(item, field))
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

        # Total column
        total_cell = ws.cell(row=row_idx, column=len(summary.rows) + 2, value=getattr(summary.totals, field))
        total_cell.number_format = fmt
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = Alignment(horizontal="right")

        # Highlight key rows
        if label in ("Net Revenue", "Gross Profit", "Net Profit"):
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # -- Column widths --
    ws.column_dimensions["A"].width = 18
    for col in range(2, len(summary.rows) + 3):
        ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = 16

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"PnL_{start_date}_{end_date}_{resolution}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
